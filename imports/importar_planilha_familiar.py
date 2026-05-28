"""Importa a aba 'Total' da planilha familiar (`despesas_Eliabe_Ana.xlsx`).

A planilha tem layout *pivotado*: cada linha é uma categoria/conta
(Moradia, Luz, Cartão de Crédito Nubank Eliabe, Ganhos Eliabe, etc.) e
cada coluna é um mês de um ano (2022-Jan, 2022-Fev, …, 2027-Dez). O
script transforma esse formato matricial em lançamentos individuais
no banco (1 por célula com valor numérico > 0).

Regras importantes:
  - **Linhas de soma/cálculo ignoradas**: 'Total Gastos', 'Défice |
    Superávit', 'Saldo', 'Poupança' (saldo, não fluxo), 'Meta',
    'Dízimos' (soma de Eliabe+Ana), 'Faculdade' (duplica
    'Faculdade - Uninter'), 'Outros' (soma da seção de despesas
    diversas — linhas 23..50 já são detalhadas).
  - **Regra anti-duplicação de cartão**: linhas
    'Cartão de Crédito - Viacredi Eliabe' e
    'Cartão de Crédito - Nubank Eliabe' são pulados quando já existe
    Fatura PDF naquela referência+conta. O detalhe do PDF é melhor que
    o agregado da planilha. O cartão da Ana não tem PDF e fica.
  - **Idempotência**: cada célula tem hash determinístico via
    `hash_lancamento_planilha`. Reimport não duplica.
  - **Data**: por não termos o dia exato, fixamos no 1º dia do mês de
    referência (`YYYY-MM-01`).

Uso CLI:
    python -m imports.importar_planilha_familiar [caminho_xlsx]

Default: `despesas_Eliabe_Ana.xlsx` na raiz do projeto.
"""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from pathlib import Path

import openpyxl

from db.models import (
    TIPO_CATEGORIA_DESPESA,
    TIPO_CATEGORIA_RECEITA,
    TIPO_LANCAMENTO_DESPESA,
    TIPO_LANCAMENTO_RECEITA,
)
from db.repository import existe_fatura_pdf, upsert_lancamento_manual

ABA_TOTAL = "Total"

MESES_PT = {
    "janeiro": 1, "fevereiro": 2, "março": 3, "marco": 3, "abril": 4,
    "maio": 5, "junho": 6, "julho": 7, "agosto": 8, "setembro": 9,
    "outubro": 10, "novembro": 11, "dezembro": 12,
}


@dataclass(frozen=True)
class LinhaConfig:
    """Como uma linha da planilha vira lançamentos no banco.

    - `descricao`: o que vira `Lancamento.descricao` (humanamente legível).
    - `categoria`: nome da `Categoria` no banco. Criada se não existir.
    - `tipo_lancamento`: `despesa` ou `receita`.
    - `tipo_categoria`: `despesa` ou `receita` (alinhar c/ tipo_lancamento).
    - `pessoa`: dono do lançamento (Eliabe, Ana ou None p/ compartilhado).
    - `conta`: conta vinculada (cartão / banco). Importante pra regra
      de dedup com PDFs (`conta_associada_pdf`).
    - `conta_associada_pdf`: quando preenchida, ativa a regra
      "pular se já existe Fatura PDF dessa conta na ref". Usar SÓ pras
      duas linhas de cartão do Eliabe.
    """

    descricao: str
    categoria: str
    tipo_lancamento: str = TIPO_LANCAMENTO_DESPESA
    tipo_categoria: str = TIPO_CATEGORIA_DESPESA
    pessoa: str | None = None
    conta: str | None = None
    conta_associada_pdf: str | None = None


# Mapeamento `nome bruto na planilha` → config. Chaves são o texto da
# coluna A, normalizado (strip + casefold) pra resistir a typos do
# planilhador. Comentários indicam a linha original do arquivo de
# referência (despesas_Eliabe_Ana.xlsx).
CONFIG: dict[str, LinhaConfig] = {
    "moradia": LinhaConfig(
        descricao="Moradia (mensal)", categoria="Moradia",
    ),
    "luz - celesc": LinhaConfig(
        descricao="Luz - Celesc", categoria="Luz",
    ),
    "água - samae": LinhaConfig(
        descricao="Água - Samae", categoria="Água",
    ),
    "internet - unifique": LinhaConfig(
        descricao="Internet - Unifique", categoria="Internet",
    ),
    "financiamento casa - caixa": LinhaConfig(
        descricao="Financiamento Casa - Caixa",
        categoria="Financiamento Casa",
        conta="Financiamento Casa - Caixa",
    ),
    "financiamento carro - bv": LinhaConfig(
        descricao="Financiamento Carro - BV",
        categoria="Financiamento Carro",
        conta="Financiamento Carro - BV",
    ),
    # Cartão Ailos do Eliabe — `Viacredi` é o nome antigo na planilha;
    # nos PDFs o cartão é "Ailos — Eliabe Gai" (com travessão em-dash).
    # Quando há PDF da referência, ignoramos a célula (PDF é fonte de
    # verdade, granular). Pessoas usam os nomes longos pra casar com
    # os já criados pelos PDFs e evitar 2 pessoas separadas (Eliabe
    # vs Eliabe Gai) na UI.
    "cartão de crédito - viacredi eliabe": LinhaConfig(
        descricao="Fatura Viacredi (mensal)",
        categoria="Cartão de Crédito",
        pessoa="Eliabe Gai",
        conta="Ailos — Eliabe Gai",
        conta_associada_pdf="Ailos — Eliabe Gai",
    ),
    "cartão de crédito - nubank eliabe": LinhaConfig(
        descricao="Fatura Nubank (mensal)",
        categoria="Cartão de Crédito",
        pessoa="Eliabe Gai",
        conta="Nubank — Eliabe Gai",
        conta_associada_pdf="Nubank — Eliabe Gai",
    ),
    "cartão de crédito - nubank ana": LinhaConfig(
        descricao="Fatura Nubank (mensal)",
        categoria="Cartão de Crédito",
        pessoa="Ana Leticia Silva Maciel",
        conta="Nubank — Ana Leticia Silva Maciel",
        conta_associada_pdf="Nubank — Ana Leticia Silva Maciel",
    ),
    "empréstimo nubank": LinhaConfig(
        descricao="Empréstimo Nubank (parcela)", categoria="Empréstimos",
    ),
    "celular tim - ana": LinhaConfig(
        descricao="Celular TIM - Ana", categoria="Celular",
        pessoa="Ana Leticia Silva Maciel",
    ),
    "dízimo - eliabe": LinhaConfig(
        descricao="Dízimo Eliabe", categoria="Dízimos", pessoa="Eliabe Gai",
    ),
    "dízimo - ana": LinhaConfig(
        descricao="Dízimo Ana", categoria="Dízimos",
        pessoa="Ana Leticia Silva Maciel",
    ),
    "faculdade - uninter - eliabe": LinhaConfig(
        descricao="Faculdade Uninter Eliabe", categoria="Educação",
        pessoa="Eliabe Gai",
    ),
    "havan": LinhaConfig(descricao="Havan", categoria="Outros Gastos"),
    "avenida - óculos": LinhaConfig(
        descricao="Avenida - Óculos", categoria="Saúde"
    ),
    "eletrecista": LinhaConfig(
        descricao="Eletricista", categoria="Outros Gastos"
    ),
    "mudança": LinhaConfig(descricao="Mudança", categoria="Outros Gastos"),
    "moto |vizinhança": LinhaConfig(
        descricao="Moto / Vizinhança", categoria="Outros Gastos"
    ),
    "renovação cnh eliabe/ana": LinhaConfig(
        descricao="Renovação CNH", categoria="Documentos"
    ),
    "manutenção carro": LinhaConfig(
        descricao="Manutenção Carro", categoria="Carro"
    ),
    "dentista ana": LinhaConfig(
        descricao="Dentista Ana", categoria="Saúde",
        pessoa="Ana Leticia Silva Maciel",
    ),
    "emprestimo nubank": LinhaConfig(
        descricao="Empréstimo Nubank (avulso)", categoria="Empréstimos"
    ),
    "ana/eliabe médico": LinhaConfig(
        descricao="Ana/Eliabe Médico", categoria="Saúde"
    ),
    "documento faculdade": LinhaConfig(
        descricao="Documento Faculdade", categoria="Educação"
    ),
    "guincho carro": LinhaConfig(
        descricao="Guincho Carro", categoria="Carro"
    ),
    "passagem onibus": LinhaConfig(
        descricao="Passagem Ônibus", categoria="Transporte"
    ),
    "feijoada/espetinho (igreja/escola)": LinhaConfig(
        descricao="Eventos Igreja/Escola", categoria="Outros Gastos"
    ),
    "petshop": LinhaConfig(
        descricao="PetShop", categoria="Outros Gastos"
    ),
    "pix pizza (clinicorp)": LinhaConfig(
        descricao="Pizza (Clinicorp)", categoria="Outros Gastos"
    ),
    "consulta sadala": LinhaConfig(
        descricao="Consulta Sadala", categoria="Saúde"
    ),
    "documento carro": LinhaConfig(
        descricao="Documento Carro", categoria="Carro"
    ),
    "iptu - casa": LinhaConfig(descricao="IPTU - Casa", categoria="Impostos"),
    "viagem": LinhaConfig(descricao="Viagem", categoria="Viagem"),
    "multa": LinhaConfig(descricao="Multa", categoria="Carro"),
    "das": LinhaConfig(descricao="DAS", categoria="Impostos"),
    "ipva": LinhaConfig(descricao="IPVA", categoria="Impostos"),
    "licenciamento": LinhaConfig(
        descricao="Licenciamento", categoria="Carro"
    ),
    "doação": LinhaConfig(descricao="Doação", categoria="Doações"),
    "emprestimo outros": LinhaConfig(
        descricao="Empréstimo Outros", categoria="Empréstimos"
    ),
    "oferta": LinhaConfig(descricao="Oferta", categoria="Doações"),
    # Receitas
    "ganhos eliabe": LinhaConfig(
        descricao="Salário Eliabe",
        categoria="Salário",
        tipo_lancamento=TIPO_LANCAMENTO_RECEITA,
        tipo_categoria=TIPO_CATEGORIA_RECEITA,
        pessoa="Eliabe Gai",
    ),
    "ganhos ana letícia": LinhaConfig(
        descricao="Salário Ana Letícia",
        categoria="Salário",
        tipo_lancamento=TIPO_LANCAMENTO_RECEITA,
        tipo_categoria=TIPO_CATEGORIA_RECEITA,
        pessoa="Ana Leticia Silva Maciel",
    ),
    "emprestimo": LinhaConfig(
        descricao="Empréstimo recebido",
        categoria="Empréstimo Recebido",
        tipo_lancamento=TIPO_LANCAMENTO_RECEITA,
        tipo_categoria=TIPO_CATEGORIA_RECEITA,
    ),
}

# Nomes que aparecem na coluna A mas devem ser pulados (totais,
# saldos, fórmulas, duplicações). Quando o mesmo nome se repete em
# linhas diferentes com semânticas opostas (caso clássico: "Outros"
# como soma de despesas vs "Outros" como receita), use
# `INDICE_OVERRIDE` abaixo pra desambiguar pela posição na planilha.
LINHAS_IGNORADAS = {
    "descrição",
    "dízimos",       # linha 17: soma de Eliabe + Ana (usamos 18/19)
    "faculdade",     # linha 20: igual à 21 'Faculdade - Uninter'
    "poupança",      # linha 54: saldo acumulado, não fluxo mensal
    "meta",          # linha 55: objetivo, não realização
    "total gastos",  # linha 56
    "défice | superávit",  # linhas 57 e 65
    "total",         # linha 64: soma de receitas
    "saldo",         # linha 66
}

# Sentinel: `INDICE_OVERRIDE[i] = None` significa "ignore essa linha
# específica mesmo que o nome esteja em CONFIG/LINHAS_IGNORADAS".
_IGNORAR = "IGNORAR_ESSA_LINHA"

# Override por número da linha (0-indexed, como `iter_rows` devolve).
# Usado pra desambiguar linhas com o mesmo nome em posições diferentes.
# Na planilha familiar atual:
#   linha 22 'Outros' → soma da seção de despesas misc (ignora)
#   linha 45 'Outros' → uma despesa individual dentro da seção
#   linha 61 'Outros' → RECEITA (subseção pós-Ganhos Ana)
#   linha 63 'Outros' → RECEITA (sub-detalhe)
INDICE_OVERRIDE: dict[int, LinhaConfig | str] = {
    22: _IGNORAR,
    45: LinhaConfig(
        descricao="Outros (gastos diversos)",
        categoria="Outros Gastos",
    ),
    61: LinhaConfig(
        descricao="Outras Receitas",
        categoria="Outras Receitas",
        tipo_lancamento=TIPO_LANCAMENTO_RECEITA,
        tipo_categoria=TIPO_CATEGORIA_RECEITA,
    ),
    63: LinhaConfig(
        descricao="Outras Receitas (avulsas)",
        categoria="Outras Receitas",
        tipo_lancamento=TIPO_LANCAMENTO_RECEITA,
        tipo_categoria=TIPO_CATEGORIA_RECEITA,
    ),
}


def _chave(texto: object) -> str:
    """Normaliza nome da linha pra casar com `CONFIG`/`LINHAS_IGNORADAS`."""
    if texto is None:
        return ""
    return str(texto).strip().lower()


def _valor_numerico(v: object) -> float | None:
    """Devolve float positivo se for número > 0; senão None.

    Cobre as 3 representações usadas no arquivo: `None`, `"-"` (hífen),
    `0`/`0.0` e números (int/float). Tudo isso vira `None` (pular)
    exceto valores numéricos > 0.
    """
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if s == "" or s == "-":
            return None
        try:
            f = float(s.replace(",", "."))
        except ValueError:
            return None
        return f if f > 0 else None
    if isinstance(v, (int, float)):
        f = float(v)
        return f if f > 0 else None
    return None


def _mapear_colunas(
    linha_ano: tuple, linha_mes: tuple
) -> dict[int, tuple[int, int]]:
    """Devolve `{col_idx: (ano, mês)}` lendo cabeçalhos.

    A linha do ano (índice 3) tem merge horizontal — o número aparece
    só na primeira coluna do bloco, então propagamos. A linha do mês
    (índice 5) tem todos os meses preenchidos.
    """
    mapa: dict[int, tuple[int, int]] = {}
    ano_atual: int | None = None
    for col_idx in range(max(len(linha_ano), len(linha_mes))):
        a = linha_ano[col_idx] if col_idx < len(linha_ano) else None
        if isinstance(a, (int, float)) and 1900 < a < 3000:
            ano_atual = int(a)
        m = linha_mes[col_idx] if col_idx < len(linha_mes) else None
        if m is None or ano_atual is None:
            continue
        nome_mes = str(m).strip().lower()
        num_mes = MESES_PT.get(nome_mes)
        if num_mes is None:
            continue
        mapa[col_idx] = (ano_atual, num_mes)
    return mapa


@dataclass
class Resultado:
    """Sumário do import — facilita teste + relatório no CLI."""

    linhas_processadas: int = 0
    linhas_ignoradas: int = 0
    linhas_desconhecidas: list[str] = None  # type: ignore[assignment]
    celulas_lidas: int = 0
    inseridos: int = 0
    duplicados: int = 0
    pulados_cartao_pdf: int = 0
    total_valor: float = 0.0

    def __post_init__(self) -> None:
        if self.linhas_desconhecidas is None:
            self.linhas_desconhecidas = []


def _ler_aba(caminho: Path) -> list[tuple]:
    """Carrega a aba 'Total' como lista de tuplas (1 por linha)."""
    if not caminho.exists():
        raise FileNotFoundError(f"Planilha não encontrada: {caminho}")
    wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
    if ABA_TOTAL not in wb.sheetnames:
        raise ValueError(
            f"Aba '{ABA_TOTAL}' não encontrada em {caminho}. "
            f"Abas disponíveis: {wb.sheetnames}"
        )
    ws = wb[ABA_TOTAL]
    return list(ws.iter_rows(values_only=True))


def migrar(caminho: str | Path = "despesas_Eliabe_Ana.xlsx") -> Resultado:
    """Roda o import; devolve `Resultado` com contadores.

    Idempotente: rodar várias vezes não duplica. A única coisa que muda
    entre rodadas é o saldo de cartões — quando você importar um PDF
    novo depois (via `gastometro`), a célula correspondente da planilha
    passa a ser pulada e a fonte oficial vira o PDF.
    """
    caminho = Path(caminho)
    rows = _ler_aba(caminho)
    if len(rows) < 7:
        raise ValueError(
            f"Planilha {caminho} tem só {len(rows)} linhas — não bate "
            f"com o layout esperado (cabeçalhos nas linhas 0..5)."
        )

    cols = _mapear_colunas(rows[3], rows[5])
    resultado = Resultado()
    arquivo_origem = caminho.name

    for linha_idx in range(6, len(rows)):
        linha = rows[linha_idx]
        if not linha:
            continue
        nome_bruto = linha[0]
        chave = _chave(nome_bruto)
        if not chave:
            continue

        # Override por índice tem prioridade — desambigua nomes
        # repetidos em posições diferentes (ex: "Outros" como soma de
        # despesas vs "Outros" como receita).
        override = INDICE_OVERRIDE.get(linha_idx)
        if override is _IGNORAR:
            resultado.linhas_ignoradas += 1
            continue
        if isinstance(override, LinhaConfig):
            conf = override
        elif chave in LINHAS_IGNORADAS:
            resultado.linhas_ignoradas += 1
            continue
        else:
            conf_opt = CONFIG.get(chave)
            if conf_opt is None:
                resultado.linhas_desconhecidas.append(str(nome_bruto))
                continue
            conf = conf_opt

        resultado.linhas_processadas += 1

        for col_idx, (ano, mes) in cols.items():
            if col_idx >= len(linha):
                continue
            valor = _valor_numerico(linha[col_idx])
            if valor is None:
                continue
            resultado.celulas_lidas += 1

            # Regra anti-duplicação de cartão: se já temos a Fatura PDF
            # dessa conta + ref, ignoramos a célula da planilha.
            if conf.conta_associada_pdf and existe_fatura_pdf(
                conta_nome=conf.conta_associada_pdf, ano=ano, mes=mes
            ):
                resultado.pulados_cartao_pdf += 1
                continue

            _, inserido = upsert_lancamento_manual(
                descricao=conf.descricao,
                valor=valor,
                ano=ano,
                mes=mes,
                categoria_nome=conf.categoria,
                pessoa_nome=conf.pessoa,
                conta_nome=conf.conta,
                tipo=conf.tipo_lancamento,
                categoria_tipo=conf.tipo_categoria,
                # Usamos `conf.descricao` (e não o nome bruto da linha)
                # como chave de hash pra evitar colisão entre linhas com
                # grafias variando — ex: `Empréstimo Nubank` (linha 15)
                # vs `Emprestimo Nubank` (linha 31) sem acento. Ambas
                # normalizam pra mesma chave, mas têm descrições
                # canônicas distintas no CONFIG.
                chave_planilha=conf.descricao,
                arquivo_origem=arquivo_origem,
            )
            if inserido:
                resultado.inseridos += 1
                resultado.total_valor += valor
            else:
                resultado.duplicados += 1

    return resultado


def _imprimir_relatorio(res: Resultado) -> None:
    print()
    print("=" * 60)
    print("  IMPORT planilha familiar — resumo")
    print("=" * 60)
    print(f"  Linhas processadas       : {res.linhas_processadas}")
    print(f"  Linhas ignoradas (total) : {res.linhas_ignoradas}")
    print(f"  Linhas desconhecidas     : {len(res.linhas_desconhecidas)}")
    print(f"  Células com valor        : {res.celulas_lidas}")
    print(f"  Lançamentos inseridos    : {res.inseridos}")
    print(f"  Duplicados (já no banco) : {res.duplicados}")
    print(f"  Cartões pulados (PDF)    : {res.pulados_cartao_pdf}")
    print(f"  Soma dos inseridos (R$)  : {res.total_valor:,.2f}")
    print("=" * 60)

    if res.linhas_desconhecidas:
        print()
        print("Atenção: linhas sem mapeamento no CONFIG (foram puladas):")
        for nome in res.linhas_desconhecidas:
            print(f"  - {nome!r}")
        print()
        print(
            "Pra incluir essas linhas, edite o dicionário CONFIG em "
            "imports/importar_planilha_familiar.py."
        )


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Importa a aba 'Total' da planilha familiar pro banco.",
    )
    parser.add_argument(
        "caminho",
        nargs="?",
        default="despesas_Eliabe_Ana.xlsx",
        help="Caminho do .xlsx (default: despesas_Eliabe_Ana.xlsx).",
    )
    args = parser.parse_args(argv)
    try:
        res = migrar(args.caminho)
    except (FileNotFoundError, ValueError) as e:
        print(f"Erro: {e}", file=sys.stderr)
        return 1
    _imprimir_relatorio(res)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
