"""Testes do `imports.importar_planilha_familiar`.

Cria planilhas sintéticas com o mesmo layout pivotado da
`despesas_Eliabe_Ana.xlsx` (aba 'Total') e valida:

- Linhas de soma/cálculo (Total Gastos, Saldo, etc.) são ignoradas.
- Cada célula numérica > 0 vira um lançamento.
- Reimport é idempotente (mesmo run não cria duplicatas).
- Regra anti-cartão-PDF: se já existe Fatura PDF dessa conta na ref,
  a célula da planilha é pulada (contada em `pulados_cartao_pdf`).
- Receitas (Ganhos Eliabe / Ana) viram `tipo='receita'`.
- Linhas sem mapeamento aparecem em `linhas_desconhecidas` em vez de
  derrubar a importação.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
import pytest


def _planilha_sintetica(path: Path) -> Path:
    """Cria um xlsx mínimo replicando o layout da planilha real.

    Layout (linhas, 0-indexed pra refletir o que `openpyxl.iter_rows`
    devolve):
      linha 0: 'ANO X', cabeçalhos de bloco
      linha 1: 'Familia ...' (ignorado)
      linha 2: vazia
      linha 3: 'Descrição', 2025 na col 1 (merge horizontal de 12 cols)
      linha 4: 'Mês/Ano' (ignorado)
      linha 5: nomes dos meses (Janeiro..Dezembro)
      linhas 6..: categorias × valores
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Total"

    # linha 1 (openpyxl=1, iter_rows=0): 'ANO 2025'
    ws.cell(row=1, column=1, value="ANO 2025")
    ws.cell(row=4, column=1, value="Descrição")
    ws.cell(row=4, column=2, value=2025)  # col B = janeiro/2025
    ws.cell(row=5, column=1, value="Mês/Ano")
    meses = [
        "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
        "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
    ]
    for i, m in enumerate(meses):
        ws.cell(row=6, column=2 + i, value=m)

    # linha 7 (idx 6): 'Internet - Unifique' — despesa, todos os meses 100.00
    ws.cell(row=7, column=1, value="Internet - Unifique")
    for i in range(12):
        ws.cell(row=7, column=2 + i, value=100.00)

    # linha 8 (idx 7): 'Luz - Celesc' — com hífens em alguns meses
    ws.cell(row=8, column=1, value="Luz - Celesc")
    ws.cell(row=8, column=2, value="-")
    ws.cell(row=8, column=3, value=85.00)
    ws.cell(row=8, column=4, value=90.00)

    # linha 9 (idx 8): 'Cartão de Crédito - Nubank Eliabe' — usado pra
    # testar regra anti-PDF.
    ws.cell(row=9, column=1, value="Cartão de Crédito - Nubank Eliabe")
    ws.cell(row=9, column=2, value=500.00)  # jan/2025 — sem PDF
    ws.cell(row=9, column=3, value=600.00)  # fev/2025 — sem PDF
    ws.cell(row=9, column=4, value=700.00)  # mar/2025 — com PDF (mock)

    # linha 10 (idx 9): 'Ganhos Eliabe' — receita
    ws.cell(row=10, column=1, value="Ganhos Eliabe")
    ws.cell(row=10, column=2, value=4000.00)
    ws.cell(row=10, column=3, value=4200.00)

    # linha 11 (idx 10): 'Total Gastos' — DEVE SER IGNORADA
    ws.cell(row=11, column=1, value="Total Gastos")
    for i in range(12):
        ws.cell(row=11, column=2 + i, value=99999.99)

    # linha 12 (idx 11): 'Saldo' — DEVE SER IGNORADA
    ws.cell(row=12, column=1, value="Saldo")
    ws.cell(row=12, column=2, value=10000.00)

    # linha 13 (idx 12): nome desconhecido — vai pra linhas_desconhecidas
    ws.cell(row=13, column=1, value="Categoria Inexistente XYZ")
    ws.cell(row=13, column=2, value=50.00)

    # linha 14 (idx 13): nome adicional desconhecido pra confirmar
    # contagem total
    ws.cell(row=14, column=1, value="Categoria Inexistente ABC")
    ws.cell(row=14, column=2, value=10.00)

    wb.save(path)
    return path


def test_planilha_sintetica_imports_corretamente(
    tmp_path: Path, banco_temporario
) -> None:
    """Roda o import e valida contagens + valores no banco."""
    from imports.importar_planilha_familiar import migrar

    xlsx = _planilha_sintetica(tmp_path / "sintetica.xlsx")
    res = migrar(xlsx)

    # Linhas mapeadas: Internet, Luz, Cartão Nubank Eliabe, Ganhos Eliabe = 4
    assert res.linhas_processadas == 4
    # Total Gastos + Saldo = 2 ignoradas
    assert res.linhas_ignoradas == 2
    # 'Categoria Inexistente XYZ' + 'Categoria Inexistente ABC' = 2 desconhecidas
    assert sorted(res.linhas_desconhecidas) == [
        "Categoria Inexistente ABC",
        "Categoria Inexistente XYZ",
    ]

    # Inseridos: 12 (Internet) + 2 (Luz, sem o '-') + 3 (Cartão) + 2 (Ganhos) = 19
    assert res.inseridos == 19
    assert res.duplicados == 0
    assert res.pulados_cartao_pdf == 0


def test_idempotencia(tmp_path: Path, banco_temporario) -> None:
    """Rodar duas vezes seguidas não duplica."""
    from imports.importar_planilha_familiar import migrar

    xlsx = _planilha_sintetica(tmp_path / "sintetica.xlsx")
    r1 = migrar(xlsx)
    r2 = migrar(xlsx)

    assert r1.inseridos > 0
    assert r2.inseridos == 0
    assert r2.duplicados == r1.inseridos


def test_regra_anti_cartao_pdf(tmp_path: Path, banco_temporario) -> None:
    """Quando já existe Fatura PDF, a célula da planilha é pulada."""
    from db.repository import upsert_fatura
    from imports.importar_planilha_familiar import migrar
    from parsers.base import Fatura, FaturaMetadata, Transacao

    fatura = Fatura(
        metadata=FaturaMetadata(
            banco="Nubank",
            titular="Eliabe Gai",
            referencia_mes="Março/2025",
            data_fechamento="01/03/2025",
            data_vencimento="10/03/2025",
            valor_total=700.0,
        ),
        transacoes=[
            Transacao(
                data="05/03/2025", descricao="Compra X", parcela="",
                cidade="Belem", valor=700.0, categoria="",
            ),
        ],
    )
    upsert_fatura(fatura, arquivo="Nubank_Mar2025.pdf")

    xlsx = _planilha_sintetica(tmp_path / "sintetica.xlsx")
    res = migrar(xlsx)

    # Mar/2025 (1 célula) é pulado porque o PDF já existe
    assert res.pulados_cartao_pdf == 1
    # Jan + Fev (2 células) do mesmo cartão entram normalmente; mar/2025 = pulado
    # Total inseridos = 19 - 1 = 18
    assert res.inseridos == 18


def test_receitas_viram_tipo_receita(
    tmp_path: Path, banco_temporario
) -> None:
    """'Ganhos Eliabe' produz lançamentos com tipo='receita'."""
    from db.repository import listar_lancamentos_df
    from imports.importar_planilha_familiar import migrar

    xlsx = _planilha_sintetica(tmp_path / "sintetica.xlsx")
    migrar(xlsx)

    df = listar_lancamentos_df()
    receitas = df[df["descricao"] == "Salário Eliabe"]
    assert len(receitas) == 2
    assert (receitas["tipo"] == "receita").all()
    assert (receitas["categoria"] == "Salário").all()
    assert (receitas["pessoa"] == "Eliabe Gai").all()


def test_celulas_zero_e_hifen_sao_puladas(
    tmp_path: Path, banco_temporario
) -> None:
    """Valores 0, 0.0, '-' e vazios não criam lançamentos."""
    from db.repository import listar_lancamentos_df
    from imports.importar_planilha_familiar import migrar

    xlsx = _planilha_sintetica(tmp_path / "sintetica.xlsx")
    migrar(xlsx)

    df = listar_lancamentos_df()
    luz = df[df["descricao"] == "Luz - Celesc"]
    # Linha 'Luz' tem '-' em jan e valores válidos em fev/mar
    assert len(luz) == 2
    assert sorted(float(v) for v in luz["valor"]) == [85.0, 90.0]


def test_data_no_primeiro_dia_do_mes(
    tmp_path: Path, banco_temporario
) -> None:
    """Sem dia exato, todos os lançamentos da planilha ficam em YYYY-MM-01."""
    from db.repository import listar_lancamentos_df
    from imports.importar_planilha_familiar import migrar

    xlsx = _planilha_sintetica(tmp_path / "sintetica.xlsx")
    migrar(xlsx)

    df = listar_lancamentos_df()
    internet = df[df["descricao"] == "Internet - Unifique"]
    assert len(internet) == 12
    # Datas: 2025-01-01, 2025-02-01, ..., 2025-12-01
    datas_esperadas = {date(2025, m, 1) for m in range(1, 13)}
    assert set(internet["data"]) == datas_esperadas


def test_planilha_real_outros_como_receita(banco_temporario) -> None:
    """Garante que as duas linhas 'Outros' depois das receitas
    (índices 61 e 63 na planilha real) entram como `tipo='receita'`
    com categoria `Outras Receitas`.

    Esse teste roda direto na planilha real do projeto (caso esteja
    presente). Skip silencioso quando o arquivo não existe — pra não
    quebrar CI em forks.
    """
    from pathlib import Path

    from db.repository import listar_lancamentos_df
    from imports.importar_planilha_familiar import migrar

    real = Path("despesas_Eliabe_Ana.xlsx")
    if not real.exists():
        pytest.skip("planilha real não encontrada no workspace")

    migrar(real)
    df = listar_lancamentos_df()
    outras = df[df["categoria"] == "Outras Receitas"]
    assert not outras.empty, "esperava lançamentos em 'Outras Receitas'"
    assert (outras["tipo"] == "receita").all()


@pytest.mark.parametrize(
    "ignorada",
    ["Descrição", "Total Gastos", "Saldo", "Faculdade", "Total", "Poupança"],
)
def test_linhas_de_soma_sao_ignoradas(
    tmp_path: Path, banco_temporario, ignorada: str
) -> None:
    """Cria planilha com só uma linha (`ignorada`) e confirma que nada entra."""
    from imports.importar_planilha_familiar import migrar

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Total"
    ws.cell(row=4, column=2, value=2025)
    meses = [
        "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
        "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
    ]
    for i, m in enumerate(meses):
        ws.cell(row=6, column=2 + i, value=m)
    ws.cell(row=7, column=1, value=ignorada)
    ws.cell(row=7, column=2, value=1234.56)
    caminho = tmp_path / f"so_{ignorada}.xlsx"
    wb.save(caminho)

    res = migrar(caminho)
    assert res.inseridos == 0
    assert res.linhas_processadas == 0
    assert res.linhas_ignoradas == 1


def test_linha_moradia_e_ignorada(tmp_path: Path, banco_temporario) -> None:
    """A linha 'Moradia' é redundante (agrega Luz/Água/Internet); ignorar."""
    import openpyxl

    from db.repository import listar_lancamentos_df
    from imports.importar_planilha_familiar import migrar

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Total"
    ws.cell(row=1, column=1, value="ANO 2025")
    ws.cell(row=4, column=1, value="Descrição")
    ws.cell(row=4, column=2, value=2025)
    ws.cell(row=5, column=1, value="Mês/Ano")
    meses = [
        "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
        "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
    ]
    for i, m in enumerate(meses):
        ws.cell(row=6, column=2 + i, value=m)
    ws.cell(row=7, column=1, value="Moradia")
    for i in range(12):
        ws.cell(row=7, column=2 + i, value=100.00)

    xlsx = tmp_path / "so_moradia.xlsx"
    wb.save(xlsx)
    res = migrar(xlsx)

    assert res.inseridos == 0
    assert res.linhas_ignoradas == 1
    df = listar_lancamentos_df()
    assert df.empty or "Moradia" not in df["descricao"].tolist()


def test_migrar_purga_descricoes_obsoletas(tmp_path: Path, banco_temporario) -> None:
    """`migrar()` apaga lançamentos legados de descrições que migraram
    pra `LINHAS_IGNORADAS` (caso clássico: 'Moradia (mensal)' que era
    inserida em versões antigas e agora é redundante)."""
    from db.models import FONTE_PLANILHA, TIPO_LANCAMENTO_DESPESA
    from db.repository import (
        listar_lancamentos_df,
        upsert_lancamento_manual,
    )
    from imports.importar_planilha_familiar import migrar

    upsert_lancamento_manual(
        descricao="Moradia (mensal)",
        valor=400.0,
        ano=2025,
        mes=5,
        categoria_nome="Moradia",
        chave_planilha="Moradia (mensal)",
        tipo=TIPO_LANCAMENTO_DESPESA,
        fonte=FONTE_PLANILHA,
    )
    upsert_lancamento_manual(
        descricao="Moradia (mensal)",
        valor=420.0,
        ano=2025,
        mes=6,
        categoria_nome="Moradia",
        chave_planilha="Moradia (mensal)",
        tipo=TIPO_LANCAMENTO_DESPESA,
        fonte=FONTE_PLANILHA,
    )

    xlsx = _planilha_sintetica(tmp_path / "sintetica.xlsx")
    res = migrar(xlsx)

    assert res.obsoletos_removidos == 2
    df = listar_lancamentos_df()
    assert (df["descricao"] == "Moradia (mensal)").sum() == 0
