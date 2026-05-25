"""
Extrator de fatura em PDF -> Excel acumulativo.

Fluxo padrão:
  1. Coloque os PDFs em `entrada/` (criada automaticamente na 1ª execução).
  2. Rode `python extrator.py`.
  3. Resultado em `saida/gastometro.xlsx` (Excel único, acumulativo entre execuções).

Faturas com o mesmo nome de arquivo já registrado no Excel são ignoradas
(para re-processar uma fatura, abra o Excel e remova a linha dela na aba
"Informações" antes de rodar novamente).

Uso:
    python extrator.py                       # processa todos os PDFs de `entrada/`
    python extrator.py Fatura.pdf            # processa um PDF específico
    python extrator.py pasta/                # processa todos os PDFs de outra pasta
    python extrator.py aprender              # salva categorias editadas no Excel
                                             # como overrides em `categorias_usuario.json`
    python extrator.py recategorizar         # re-aplica `categorizar()` em todas as
                                             # transações do Excel (sem reler PDFs),
                                             # reconstrói as abas analíticas e
                                             # preserva filtros/tabelas existentes
"""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from categorias import (
    CATEGORIAS_USUARIO_ARQUIVO,
    categorizar,
    categorizar_pelo_dicionario,
    salvar_categorias_usuario,
)
from parsers import Fatura, extrair_fatura
from parsers.base import MES_POR_NUMERO


RAIZ = Path(__file__).parent
PASTA_ENTRADA = RAIZ / "entrada"
PASTA_SAIDA = RAIZ / "saida"
ARQUIVO_SAIDA = "gastometro.xlsx"

COLUNAS_INFO = [
    "Arquivo",
    "Banco",
    "Titular",
    "Cartão",
    "Referência",
    "Data de fechamento",
    "Data de vencimento",
    "Valor total (R$)",
    "Qtde. de transações",
]

COLUNAS_TRANSACOES = [
    "Arquivo",
    "Banco",
    "Titular",
    "Cartão",
    "Referência",
    "Data",
    "Descrição",
    "Parcela",
    "Cidade",
    "Valor (R$)",
    "Categoria",
]

MES_POR_NOME = {nome: num for num, nome in MES_POR_NUMERO.items()}

TOLERANCIA_TOTAL = 0.01

ABAS_COM_FILTRO = {
    "Informações",
    "Transações",
    "Top Comerciantes",
    "Recorrentes",
    "Maiores Gastos",
}


def _identificador_cartao(banco: str, titular: str) -> str:
    """Combina banco e titular num rótulo único para distinguir cartões.

    Útil quando o mesmo banco tem cartões de titulares diferentes (ex.: Nubank
    do João vs Nubank da Maria) ou quando o mesmo titular tem cartões em vários
    bancos. Formato: `Banco — Titular` (em branco no titular cai em `Banco`).
    """
    banco = (banco or "").strip()
    titular = (titular or "").strip()
    if not banco and not titular:
        return ""
    if not titular:
        return banco
    if not banco:
        return titular
    return f"{banco} — {titular}"


def _conciliar_total(fatura: Fatura) -> None:
    """Compara o total extraído do PDF com a soma das transações.

    - Se o parser não conseguiu extrair o total (`valor_total == 0`),
      preenche com a soma e avisa que estamos confiando só na soma.
    - Se há total extraído mas ele diverge da soma além de `TOLERANCIA_TOTAL`,
      imprime um aviso. Quando a diferença bate com a soma dos estornos
      (valores negativos no extrato), o aviso é informativo: o banco
      tipicamente computa o total bruto e os estornos aparecem como
      crédito separado.
    - Caso contrário, fica em silêncio.
    """
    meta = fatura.metadata
    soma = sum(t.valor for t in fatura.transacoes)
    if meta.valor_total == 0:
        if fatura.transacoes:
            meta.valor_total = soma
            print(
                f"  AVISO: total da fatura não foi extraído do PDF; "
                f"usando soma das transações (R$ {soma:.2f})."
            )
        return
    diferenca = meta.valor_total - soma
    if abs(diferenca) <= TOLERANCIA_TOTAL:
        return

    soma_estornos = sum(t.valor for t in fatura.transacoes if t.valor < 0)
    sinal = "+" if diferenca > 0 else "-"
    cabecalho = (
        f"  AVISO: total da fatura R$ {meta.valor_total:.2f} difere da soma "
        f"das transações R$ {soma:.2f} ({sinal}R$ {abs(diferenca):.2f})."
    )
    if soma_estornos < 0 and abs(diferenca + soma_estornos) <= TOLERANCIA_TOTAL:
        print(
            f"{cabecalho} Diferença equivale aos estornos detectados "
            f"(R$ {soma_estornos:.2f}); o banco computa o total bruto."
        )
    else:
        print(f"{cabecalho} Pode haver lançamento não capturado pelo parser.")


def _fatura_para_dicts(fatura: Fatura, arquivo: str) -> tuple[dict, list[dict]]:
    meta = fatura.metadata
    cartao = _identificador_cartao(meta.banco, meta.titular)
    info = {
        "Arquivo": arquivo,
        "Banco": meta.banco,
        "Titular": meta.titular,
        "Cartão": cartao,
        "Referência": meta.referencia_mes,
        "Data de fechamento": meta.data_fechamento,
        "Data de vencimento": meta.data_vencimento,
        "Valor total (R$)": meta.valor_total,
        "Qtde. de transações": len(fatura.transacoes),
    }
    linhas = [
        {
            "Arquivo": arquivo,
            "Banco": meta.banco,
            "Titular": meta.titular,
            "Cartão": cartao,
            "Referência": meta.referencia_mes,
            "Data": t.data,
            "Descrição": t.descricao,
            "Parcela": t.parcela,
            "Cidade": t.cidade,
            "Valor (R$)": t.valor,
            "Categoria": t.categoria,
        }
        for t in fatura.transacoes
    ]
    return info, linhas


def _garantir_coluna_cartao(df: pd.DataFrame) -> pd.DataFrame:
    """Calcula a coluna `Cartão` a partir de `Banco`/`Titular` quando ausente.

    Retrocompatibilidade: Excels gerados antes desta coluna existir são
    enriquecidos ao serem carregados, evitando que o usuário precise
    reprocessar tudo do zero.
    """
    if df.empty:
        return df
    if "Cartão" in df.columns and df["Cartão"].astype(str).str.strip().ne("").any():
        return df
    banco = df.get("Banco", pd.Series([""] * len(df))).astype(str)
    titular = df.get("Titular", pd.Series([""] * len(df))).astype(str)
    df = df.copy()
    df["Cartão"] = [_identificador_cartao(b, t) for b, t in zip(banco, titular)]
    return df


def _carregar_excel_existente(
    caminho: Path,
) -> tuple[pd.DataFrame, pd.DataFrame, bool]:
    """Lê informações e transações de um Excel acumulativo existente.

    Retorna DataFrames vazios (com as colunas esperadas) se o arquivo
    ainda não existir ou se as abas estiverem incompatíveis. O terceiro
    valor é `True` quando alguma migração de schema foi necessária
    (ex.: Excel antigo sem coluna `Cartão`), indicando que o arquivo
    precisa ser regravado mesmo sem faturas novas.
    """
    df_info = pd.DataFrame(columns=COLUNAS_INFO)
    df_transacoes = pd.DataFrame(columns=COLUNAS_TRANSACOES)
    precisou_migrar = False
    if not caminho.exists():
        return df_info, df_transacoes, precisou_migrar
    try:
        df_info_lido = pd.read_excel(caminho, sheet_name="Informações")
        if "Arquivo" in df_info_lido.columns:
            if "Cartão" not in df_info_lido.columns:
                precisou_migrar = True
            df_info = _garantir_coluna_cartao(df_info_lido)
    except Exception:
        pass
    try:
        df_t_lido = pd.read_excel(caminho, sheet_name="Transações")
        if "Arquivo" in df_t_lido.columns:
            if "Cartão" not in df_t_lido.columns:
                precisou_migrar = True
            df_transacoes = _garantir_coluna_cartao(df_t_lido)
    except Exception:
        pass
    return df_info, df_transacoes, precisou_migrar


def _ordenar_transacoes(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    df = df.copy()
    df["_data_ord"] = pd.to_datetime(df["Data"], format="%d/%m/%Y", errors="coerce")
    df = df.sort_values(["_data_ord", "Arquivo"], kind="stable").reset_index(drop=True)
    return df.drop(columns="_data_ord")


def _construir_resumo_geral(df_transacoes: pd.DataFrame) -> pd.DataFrame:
    if df_transacoes.empty:
        return pd.DataFrame(columns=["Categoria", "Valor (R$)"])
    resumo = (
        df_transacoes.groupby("Categoria", as_index=False)["Valor (R$)"]
        .sum()
        .sort_values("Valor (R$)", ascending=False)
        .reset_index(drop=True)
    )
    total = pd.DataFrame(
        [{"Categoria": "TOTAL GERAL", "Valor (R$)": resumo["Valor (R$)"].sum()}]
    )
    return pd.concat([resumo, total], ignore_index=True)


def _chave_ordenacao_referencia(referencia: str) -> tuple[int, int]:
    """Converte 'Maio/2026' em (2026, 5) para ordenar cronologicamente."""
    try:
        nome_mes, ano = referencia.split("/")
        return (int(ano), MES_POR_NOME.get(nome_mes, 0))
    except (ValueError, AttributeError):
        return (0, 0)


def _construir_resumo_mensal(df_transacoes: pd.DataFrame) -> pd.DataFrame:
    """Pivot Referência × Categoria, com Total mensal e variação % vs mês anterior.

    A coluna `Variação %` está em pontos percentuais (12.3 = +12,3%) e fica
    em branco no 1º mês e na linha `TOTAL`. Usada também como base do
    `BarChart` (coluna `Total`) inserido na aba.
    """
    if df_transacoes.empty or "Referência" not in df_transacoes.columns:
        return pd.DataFrame()
    pivot = pd.pivot_table(
        df_transacoes,
        index="Referência",
        columns="Categoria",
        values="Valor (R$)",
        aggfunc="sum",
        fill_value=0.0,
    )
    referencias_ordenadas = sorted(pivot.index, key=_chave_ordenacao_referencia)
    pivot = pivot.loc[referencias_ordenadas]
    pivot["Total"] = pivot.sum(axis=1)

    totais = pivot["Total"].tolist()
    variacoes: list[float | None] = []
    for i, atual in enumerate(totais):
        if i == 0:
            variacoes.append(None)
            continue
        anterior = totais[i - 1]
        if anterior == 0:
            variacoes.append(None)
        else:
            variacoes.append((atual - anterior) / anterior * 100.0)
    pivot["Variação %"] = variacoes

    soma_linhas = pivot.drop(columns=["Variação %"]).sum(axis=0)
    pivot.loc["TOTAL"] = soma_linhas
    pivot.at["TOTAL", "Variação %"] = None
    return pivot.reset_index()


def _construir_cartao_x_mes(df_transacoes: pd.DataFrame) -> pd.DataFrame:
    """Pivot referência (linhas) × cartão (colunas), com totais.

    Responde "quanto cada cartão gastou em cada mês". A coluna `Total`
    soma todos os cartões no mês, e a linha `TOTAL` soma cada cartão no
    período inteiro.
    """
    if df_transacoes.empty:
        return pd.DataFrame()
    if "Referência" not in df_transacoes.columns or "Cartão" not in df_transacoes.columns:
        return pd.DataFrame()
    pivot = pd.pivot_table(
        df_transacoes,
        index="Referência",
        columns="Cartão",
        values="Valor (R$)",
        aggfunc="sum",
        fill_value=0.0,
    )
    referencias_ordenadas = sorted(pivot.index, key=_chave_ordenacao_referencia)
    pivot = pivot.loc[referencias_ordenadas]
    pivot["Total"] = pivot.sum(axis=1)
    pivot.loc["TOTAL"] = pivot.sum(axis=0)
    return pivot.reset_index()


def _construir_cartao_x_categoria(df_transacoes: pd.DataFrame) -> pd.DataFrame:
    """Pivot categoria (linhas) × cartão (colunas), com totais.

    Responde "como cada cartão se distribui por categoria". Útil para
    decidir qual cartão usar para qual tipo de gasto, ou perceber
    desbalanceamentos (ex.: um cartão concentrando Mercado).
    """
    if df_transacoes.empty:
        return pd.DataFrame()
    if "Categoria" not in df_transacoes.columns or "Cartão" not in df_transacoes.columns:
        return pd.DataFrame()
    pivot = pd.pivot_table(
        df_transacoes,
        index="Categoria",
        columns="Cartão",
        values="Valor (R$)",
        aggfunc="sum",
        fill_value=0.0,
    )
    pivot["Total"] = pivot.sum(axis=1)
    pivot = pivot.sort_values("Total", ascending=False)
    pivot.loc["TOTAL"] = pivot.sum(axis=0)
    return pivot.reset_index()


def _construir_resumo_por_cartao(
    df_info: pd.DataFrame, df_transacoes: pd.DataFrame
) -> pd.DataFrame:
    """Uma linha por cartão (Banco + Titular) com agregados úteis.

    Colunas:
      - Cartão, Banco, Titular
      - Qtde. Faturas, Qtde. Transações
      - Primeira / Última Referência (cobertura temporal)
      - Total Gasto (R$) — soma das transações (inclui estornos)
      - Média por Fatura (R$), Ticket Médio (R$)
    """
    if df_transacoes.empty or "Cartão" not in df_transacoes.columns:
        return pd.DataFrame()

    df_t = df_transacoes.copy()
    df_t["_ref_ord"] = df_t["Referência"].apply(_chave_ordenacao_referencia)

    linhas: list[dict] = []
    for cartao, grupo in df_t.groupby("Cartão", sort=False):
        if not cartao:
            continue
        banco = str(grupo["Banco"].iloc[0]) if "Banco" in grupo.columns else ""
        titular = str(grupo["Titular"].iloc[0]) if "Titular" in grupo.columns else ""
        refs = sorted(set(grupo["Referência"]), key=_chave_ordenacao_referencia)
        total = float(grupo["Valor (R$)"].sum())
        qtd_tx = len(grupo)
        qtd_faturas = grupo["Arquivo"].nunique() if "Arquivo" in grupo.columns else len(refs)
        media_fatura = total / qtd_faturas if qtd_faturas else 0.0
        ticket = total / qtd_tx if qtd_tx else 0.0
        linhas.append(
            {
                "Cartão": cartao,
                "Banco": banco,
                "Titular": titular,
                "Qtde. Faturas": qtd_faturas,
                "Qtde. Transações": qtd_tx,
                "Primeira Referência": refs[0] if refs else "",
                "Última Referência": refs[-1] if refs else "",
                "Total Gasto (R$)": total,
                "Média por Fatura (R$)": media_fatura,
                "Ticket Médio (R$)": ticket,
            }
        )

    if not linhas:
        return pd.DataFrame()

    df_resumo = pd.DataFrame(linhas).sort_values(
        "Total Gasto (R$)", ascending=False
    )

    total_geral = {
        "Cartão": "TOTAL",
        "Banco": "",
        "Titular": "",
        "Qtde. Faturas": int(df_resumo["Qtde. Faturas"].sum()),
        "Qtde. Transações": int(df_resumo["Qtde. Transações"].sum()),
        "Primeira Referência": "",
        "Última Referência": "",
        "Total Gasto (R$)": float(df_resumo["Total Gasto (R$)"].sum()),
        "Média por Fatura (R$)": (
            float(df_resumo["Total Gasto (R$)"].sum())
            / int(df_resumo["Qtde. Faturas"].sum())
            if df_resumo["Qtde. Faturas"].sum()
            else 0.0
        ),
        "Ticket Médio (R$)": (
            float(df_resumo["Total Gasto (R$)"].sum())
            / int(df_resumo["Qtde. Transações"].sum())
            if df_resumo["Qtde. Transações"].sum()
            else 0.0
        ),
    }
    return pd.concat(
        [df_resumo, pd.DataFrame([total_geral])], ignore_index=True
    )


def _construir_top_transacoes(
    df_transacoes: pd.DataFrame, top_n: int = 20
) -> pd.DataFrame:
    """Top N transações individuais por valor (apenas gastos, ignora estornos).

    Diferente de `Top Comerciantes` (que agrupa por descrição), aqui cada
    linha é uma compra isolada — útil para revisar os maiores impactos
    pontuais no período (parcela cara, hotel, eletrodoméstico, etc.).
    """
    if df_transacoes.empty:
        return pd.DataFrame()

    gastos = df_transacoes[df_transacoes["Valor (R$)"] > 0].copy()
    if gastos.empty:
        return pd.DataFrame()

    colunas = [
        "Data",
        "Referência",
        "Descrição",
        "Categoria",
        "Cartão",
        "Parcela",
        "Cidade",
        "Valor (R$)",
    ]
    colunas = [c for c in colunas if c in gastos.columns]
    return (
        gastos.sort_values("Valor (R$)", ascending=False)
        .head(top_n)[colunas]
        .reset_index(drop=True)
    )


def _construir_top_comerciantes(
    df_transacoes: pd.DataFrame, top_n: int = 30
) -> pd.DataFrame:
    """Top descrições por valor acumulado (apenas gastos, ignora estornos).

    Útil para identificar onde o dinheiro está indo de fato, independente
    de como foi categorizado. Mostra também ticket médio e cartão(s) onde
    apareceu, ajudando a perceber duplicidade entre cartões.
    """
    if df_transacoes.empty or "Descrição" not in df_transacoes.columns:
        return pd.DataFrame()

    gastos = df_transacoes[df_transacoes["Valor (R$)"] > 0].copy()
    if gastos.empty:
        return pd.DataFrame()

    agg = (
        gastos.groupby("Descrição")
        .agg(
            **{
                "Qtde.": ("Valor (R$)", "count"),
                "Total (R$)": ("Valor (R$)", "sum"),
                "Ticket Médio (R$)": ("Valor (R$)", "mean"),
                "Categoria": ("Categoria", lambda s: s.mode().iat[0] if not s.mode().empty else ""),
                "Cartão(ões)": ("Cartão", lambda s: ", ".join(sorted({str(v) for v in s if str(v).strip()}))),
            }
        )
        .reset_index()
        .sort_values("Total (R$)", ascending=False)
        .head(top_n)
        .reset_index(drop=True)
    )
    return agg[
        [
            "Descrição",
            "Categoria",
            "Cartão(ões)",
            "Qtde.",
            "Total (R$)",
            "Ticket Médio (R$)",
        ]
    ]


def _construir_recorrentes(
    df_transacoes: pd.DataFrame, meses_min: int = 3
) -> pd.DataFrame:
    """Descrições que aparecem em pelo menos `meses_min` meses distintos.

    Excelente para mapear gastos fixos (assinaturas, mensalidades) que
    pesam todo mês mesmo sendo individualmente pequenos. Mostra média
    mensal real (dividida pelo nº de meses em que apareceu).
    """
    if df_transacoes.empty or "Descrição" not in df_transacoes.columns:
        return pd.DataFrame()
    if "Referência" not in df_transacoes.columns:
        return pd.DataFrame()

    gastos = df_transacoes[df_transacoes["Valor (R$)"] > 0].copy()
    if gastos.empty:
        return pd.DataFrame()

    agg = (
        gastos.groupby("Descrição")
        .agg(
            **{
                "Meses": ("Referência", "nunique"),
                "Qtde. Transações": ("Valor (R$)", "count"),
                "Total (R$)": ("Valor (R$)", "sum"),
                "Média Mensal (R$)": ("Valor (R$)", "sum"),
                "Categoria": ("Categoria", lambda s: s.mode().iat[0] if not s.mode().empty else ""),
                "Cartão(ões)": ("Cartão", lambda s: ", ".join(sorted({str(v) for v in s if str(v).strip()}))),
            }
        )
        .reset_index()
    )
    agg = agg[agg["Meses"] >= meses_min].copy()
    if agg.empty:
        return pd.DataFrame()
    agg["Média Mensal (R$)"] = agg["Total (R$)"] / agg["Meses"]
    agg = agg.sort_values(
        ["Total (R$)", "Meses"], ascending=[False, False]
    ).reset_index(drop=True)
    return agg[
        [
            "Descrição",
            "Categoria",
            "Cartão(ões)",
            "Meses",
            "Qtde. Transações",
            "Total (R$)",
            "Média Mensal (R$)",
        ]
    ]


def salvar_excel_acumulativo(
    df_info: pd.DataFrame, df_transacoes: pd.DataFrame, destino: Path
) -> None:
    df_transacoes = _ordenar_transacoes(df_transacoes)
    df_info = df_info.reindex(columns=COLUNAS_INFO)

    df_resumo = _construir_resumo_geral(df_transacoes)
    df_mensal = _construir_resumo_mensal(df_transacoes)
    df_cartao_mes = _construir_cartao_x_mes(df_transacoes)
    df_cartao_cat = _construir_cartao_x_categoria(df_transacoes)
    df_resumo_cartao = _construir_resumo_por_cartao(df_info, df_transacoes)
    df_maiores = _construir_top_transacoes(df_transacoes)
    df_top = _construir_top_comerciantes(df_transacoes)
    df_recorrentes = _construir_recorrentes(df_transacoes)

    destino.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(destino, engine="openpyxl") as writer:
        df_info.to_excel(writer, sheet_name="Informações", index=False)
        df_transacoes.to_excel(writer, sheet_name="Transações", index=False)
        df_resumo.to_excel(writer, sheet_name="Resumo por Categoria", index=False)
        if not df_mensal.empty:
            df_mensal.to_excel(writer, sheet_name="Resumo Mensal", index=False)
        if not df_resumo_cartao.empty:
            df_resumo_cartao.to_excel(
                writer, sheet_name="Resumo por Cartão", index=False
            )
        if not df_cartao_mes.empty:
            df_cartao_mes.to_excel(writer, sheet_name="Cartão x Mês", index=False)
        if not df_cartao_cat.empty:
            df_cartao_cat.to_excel(
                writer, sheet_name="Cartão x Categoria", index=False
            )
        if not df_maiores.empty:
            df_maiores.to_excel(writer, sheet_name="Maiores Gastos", index=False)
        if not df_top.empty:
            df_top.to_excel(writer, sheet_name="Top Comerciantes", index=False)
        if not df_recorrentes.empty:
            df_recorrentes.to_excel(writer, sheet_name="Recorrentes", index=False)

        _formatar_planilha(writer, "Informações", df_info)
        _formatar_planilha(writer, "Transações", df_transacoes)
        _formatar_planilha(writer, "Resumo por Categoria", df_resumo)
        if not df_mensal.empty:
            _formatar_planilha(writer, "Resumo Mensal", df_mensal)
        if not df_resumo_cartao.empty:
            _formatar_planilha(writer, "Resumo por Cartão", df_resumo_cartao)
        if not df_cartao_mes.empty:
            _formatar_planilha(writer, "Cartão x Mês", df_cartao_mes)
        if not df_cartao_cat.empty:
            _formatar_planilha(writer, "Cartão x Categoria", df_cartao_cat)
        if not df_maiores.empty:
            _formatar_planilha(writer, "Maiores Gastos", df_maiores)
        if not df_top.empty:
            _formatar_planilha(writer, "Top Comerciantes", df_top)
        if not df_recorrentes.empty:
            _formatar_planilha(writer, "Recorrentes", df_recorrentes)

        _adicionar_grafico_categorias(writer, df_resumo)
        if not df_mensal.empty:
            _adicionar_grafico_mensal(writer, df_mensal)


ABAS_VALOR_PIVOT = {"Resumo Mensal", "Cartão x Mês", "Cartão x Categoria"}


def _formatar_planilha(writer, nome_aba: str, df: pd.DataFrame) -> None:
    aba = writer.sheets[nome_aba]

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx in range(1, len(df.columns) + 1):
        cell = aba.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    colunas_percent_idx = [
        df.columns.get_loc(c) + 1 for c in df.columns if "%" in str(c)
    ]
    if nome_aba in ABAS_VALOR_PIVOT:
        colunas_valor_idx = [
            i for i in range(2, len(df.columns) + 1) if i not in colunas_percent_idx
        ]
    else:
        colunas_valor_idx = [
            df.columns.get_loc(c) + 1
            for c in df.columns
            if (
                "Valor" in c
                or "Total" in c
                or "Médio" in c
                or "Média" in c
                or c == "Ticket"
            )
            and "%" not in str(c)
        ]
    for col_idx in colunas_valor_idx:
        for row in range(2, len(df) + 2):
            celula = aba.cell(row=row, column=col_idx)
            if isinstance(celula.value, (int, float)):
                celula.number_format = "R$ #,##0.00"
    for col_idx in colunas_percent_idx:
        for row in range(2, len(df) + 2):
            celula = aba.cell(row=row, column=col_idx)
            if isinstance(celula.value, (int, float)):
                celula.number_format = '+0.0"%";-0.0"%";0.0"%"'

    for col_idx, coluna in enumerate(df.columns, start=1):
        valores = [str(v) for v in df[coluna].astype(str).tolist()] + [str(coluna)]
        largura = min(max(len(v) for v in valores) + 2, 60)
        aba.column_dimensions[get_column_letter(col_idx)].width = largura

    if nome_aba in ABAS_COM_FILTRO and len(df) > 0 and len(df.columns) > 0:
        ultima_col = get_column_letter(len(df.columns))
        aba.auto_filter.ref = f"A1:{ultima_col}{len(df) + 1}"

    aba.freeze_panes = "A2"


def _adicionar_grafico_categorias(writer, df_resumo: pd.DataFrame) -> None:
    """Insere `PieChart` na aba 'Resumo por Categoria' (exclui linha TOTAL GERAL).

    Resumo é uma coluna `Categoria` (texto) + coluna `Valor (R$)`. As
    categorias começam na linha 2 e a última é `TOTAL GERAL`, então
    pegamos até `len(df_resumo)` (a antepenúltima linha de dados +
    cabeçalho), e o gráfico fica posicionado em D2.
    """
    if df_resumo.empty or len(df_resumo) < 3:
        return
    aba = writer.sheets.get("Resumo por Categoria")
    if aba is None:
        return
    ultima_linha_dados = len(df_resumo)
    chart = PieChart()
    chart.title = "Distribuição por Categoria"
    chart.height = 12
    chart.width = 18
    rotulos = Reference(aba, min_col=1, min_row=2, max_row=ultima_linha_dados)
    dados = Reference(
        aba, min_col=2, min_row=1, max_row=ultima_linha_dados
    )
    chart.add_data(dados, titles_from_data=True)
    chart.set_categories(rotulos)
    aba.add_chart(chart, "D2")


def _adicionar_grafico_mensal(writer, df_mensal: pd.DataFrame) -> None:
    """Insere `BarChart` na aba 'Resumo Mensal' usando a coluna `Total`.

    Ignora a linha `TOTAL` (última) no eixo X. Posicionado embaixo dos
    dados, com 2 linhas de respiro.
    """
    if df_mensal.empty or "Total" not in df_mensal.columns:
        return
    aba = writer.sheets.get("Resumo Mensal")
    if aba is None:
        return
    ultima_linha_dados = len(df_mensal)
    n_meses = ultima_linha_dados - 1
    if n_meses < 2:
        return
    col_total = df_mensal.columns.get_loc("Total") + 1

    chart = BarChart()
    chart.type = "col"
    chart.title = "Total Mensal"
    chart.y_axis.title = "R$"
    chart.x_axis.title = "Mês"
    chart.height = 10
    chart.width = 22
    dados = Reference(aba, min_col=col_total, min_row=1, max_row=1 + n_meses)
    categorias = Reference(aba, min_col=1, min_row=2, max_row=1 + n_meses)
    chart.add_data(dados, titles_from_data=True)
    chart.set_categories(categorias)
    aba.add_chart(chart, f"A{ultima_linha_dados + 3}")


def _descobrir_pdfs(alvo: Path | None) -> list[Path]:
    if alvo is None:
        PASTA_ENTRADA.mkdir(parents=True, exist_ok=True)
        PASTA_SAIDA.mkdir(parents=True, exist_ok=True)
        pdfs = sorted(PASTA_ENTRADA.glob("*.pdf"))
        if not pdfs:
            print(f"Nenhum PDF encontrado em '{PASTA_ENTRADA.name}/'.")
            print(
                f"Coloque os PDFs em '{PASTA_ENTRADA.name}/' e rode novamente, "
                f"ou passe um caminho como argumento."
            )
        return pdfs
    if alvo.is_file() and alvo.suffix.lower() == ".pdf":
        return [alvo]
    if alvo.is_dir():
        return sorted(alvo.glob("*.pdf"))
    print(f"Caminho inválido: {alvo}")
    return []


def processar(pdfs: Iterable[Path], destino: Path) -> None:
    df_info_existente, df_transacoes_existente, precisou_migrar = (
        _carregar_excel_existente(destino)
    )
    arquivos_no_excel = set(df_info_existente["Arquivo"].astype(str).tolist())
    if precisou_migrar:
        print(
            "Excel existente em formato antigo (sem coluna 'Cartão'). "
            "Será regravado com as novas abas analíticas."
        )

    novos_info: list[dict] = []
    novas_transacoes: list[dict] = []
    ignorados: list[str] = []
    pulados_sem_transacao: list[str] = []

    for pdf in pdfs:
        if pdf.name in arquivos_no_excel:
            ignorados.append(pdf.name)
            print(f"\nIgnorado (já no Excel): {pdf.name}")
            continue

        print(f"\nProcessando: {pdf.name}")
        try:
            fatura = extrair_fatura(pdf)
        except Exception as exc:
            print(f"  Erro ao ler {pdf.name}: {exc}")
            continue

        meta = fatura.metadata
        print(
            f"  Banco: {meta.banco} | Titular: {meta.titular or '—'} | "
            f"Referência: {meta.referencia_mes or '—'}"
        )
        print(
            f"  Fechamento: {meta.data_fechamento or '—'} | "
            f"Vencimento: {meta.data_vencimento or '—'}"
        )
        print(f"  {len(fatura.transacoes)} transações encontradas.")

        if not fatura.transacoes:
            pulados_sem_transacao.append(pdf.name)
            continue

        _conciliar_total(fatura)
        info, linhas = _fatura_para_dicts(fatura, pdf.name)
        novos_info.append(info)
        novas_transacoes.extend(linhas)

    if not novos_info:
        if precisou_migrar and not df_info_existente.empty:
            salvar_excel_acumulativo(
                df_info_existente, df_transacoes_existente, destino
            )
            print(
                f"\nExcel migrado para o novo formato: {destino}\n"
                f"  Total no arquivo: {len(df_info_existente)} faturas, "
                f"{len(df_transacoes_existente)} transações."
            )
            _imprimir_top_outros_gastos(df_transacoes_existente)
            _imprimir_comparativo_mensal(df_transacoes_existente)
            return
        print("\nNenhuma fatura nova para adicionar ao Excel.")
        if ignorados:
            print(f"  Ignoradas (já no Excel): {len(ignorados)}")
        if pulados_sem_transacao:
            print(f"  Sem transações: {len(pulados_sem_transacao)}")
        return

    df_info_final = pd.concat(
        [df_info_existente, pd.DataFrame(novos_info)], ignore_index=True
    )
    df_transacoes_final = pd.concat(
        [df_transacoes_existente, pd.DataFrame(novas_transacoes)], ignore_index=True
    )
    salvar_excel_acumulativo(df_info_final, df_transacoes_final, destino)

    print(f"\nExcel acumulativo atualizado: {destino}")
    print(
        f"  Total no arquivo: {len(df_info_final)} faturas, "
        f"{len(df_transacoes_final)} transações."
    )
    print(f"  Faturas adicionadas nesta execução: {len(novos_info)}.")
    if ignorados:
        print(f"  Ignoradas (já no Excel): {len(ignorados)}.")

    _imprimir_top_outros_gastos(df_transacoes_final)
    _imprimir_comparativo_mensal(df_transacoes_final)


def _formatar_brl(valor: float) -> str:
    """Formata em moeda BR sem depender de locale: `1.234,56`."""
    s = f"{abs(valor):,.2f}"
    s = s.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"-R$ {s}" if valor < 0 else f"R$ {s}"


def _imprimir_comparativo_mensal(df_transacoes: pd.DataFrame) -> None:
    """Compara último mês com o anterior: total geral + categorias com
    maior variação absoluta. Imprime nada se houver menos de 2 meses.
    """
    if df_transacoes is None or df_transacoes.empty:
        return
    if "Referência" not in df_transacoes.columns:
        return
    if "Categoria" not in df_transacoes.columns:
        return

    refs = [r for r in df_transacoes["Referência"].astype(str).unique() if r]
    refs.sort(key=_chave_ordenacao_referencia)
    if len(refs) < 2:
        return

    ultimo, anterior = refs[-1], refs[-2]

    def soma_por_categoria(ref: str) -> dict[str, float]:
        sub = df_transacoes[df_transacoes["Referência"] == ref]
        return (
            sub.groupby("Categoria")["Valor (R$)"]
            .sum()
            .to_dict()
        )

    cats_ult = soma_por_categoria(ultimo)
    cats_ant = soma_por_categoria(anterior)
    total_ult = float(sum(cats_ult.values()))
    total_ant = float(sum(cats_ant.values()))

    print(f"\nComparativo: {ultimo} vs {anterior}")
    if total_ant:
        var_pct = (total_ult - total_ant) / total_ant * 100.0
        sinal = "+" if var_pct >= 0 else ""
        diff = total_ult - total_ant
        sinal_abs = "+" if diff >= 0 else "-"
        print(
            f"  TOTAL                 {_formatar_brl(total_ult):>14}  "
            f"({sinal}{var_pct:.1f}% / {sinal_abs}{_formatar_brl(abs(diff))[3:]}"
            f" vs {_formatar_brl(total_ant)})"
        )
    else:
        print(
            f"  TOTAL                 {_formatar_brl(total_ult):>14}  "
            f"(mês anterior R$ 0,00)"
        )

    todas = set(cats_ult) | set(cats_ant)
    diffs: list[tuple[str, float, float, float]] = []
    for cat in todas:
        a = float(cats_ult.get(cat, 0.0))
        b = float(cats_ant.get(cat, 0.0))
        diffs.append((cat, a, b, a - b))
    diffs.sort(key=lambda t: abs(t[3]), reverse=True)

    for cat, atual, ant, diff in diffs[:8]:
        if abs(diff) < 0.01:
            continue
        sinal_abs = "+" if diff >= 0 else "-"
        valor_atual = _formatar_brl(atual)
        if ant == 0:
            print(
                f"  {cat:21}  {valor_atual:>14}  "
                f"(novo, {sinal_abs}{_formatar_brl(abs(diff))[3:]})"
            )
        else:
            var_pct = diff / ant * 100.0
            sinal = "+" if var_pct >= 0 else ""
            print(
                f"  {cat:21}  {valor_atual:>14}  "
                f"({sinal}{var_pct:.1f}% / {sinal_abs}{_formatar_brl(abs(diff))[3:]})"
            )


def _imprimir_top_outros_gastos(df_transacoes: pd.DataFrame, top_n: int = 10) -> None:
    """Imprime as descrições mais frequentes que caíram em `Outros Gastos`,
    ordenadas pelo valor total. Ajuda a evoluir o dicionário/overrides."""
    if df_transacoes is None or df_transacoes.empty:
        return
    if "Categoria" not in df_transacoes.columns or "Descrição" not in df_transacoes.columns:
        return

    outros = df_transacoes[
        (df_transacoes["Categoria"] == "Outros Gastos")
        & (df_transacoes["Valor (R$)"] > 0)
    ]
    if outros.empty:
        return

    agregado = (
        outros.groupby("Descrição")["Valor (R$)"]
        .agg(soma="sum", n="count")
        .reset_index()
        .sort_values("soma", ascending=False)
        .head(top_n)
    )

    print(
        f"\nTop {len(agregado)} descrições em 'Outros Gastos' "
        f"(acumulado no Excel):"
    )
    for _, linha in agregado.iterrows():
        print(
            f"  R$ {linha['soma']:9.2f}  ({int(linha['n']):2d}x)  "
            f"{linha['Descrição']}"
        )
    print(
        "  Para categorizar: edite `categorias.py` (regra geral) ou rode "
        "`python extrator.py aprender` depois de ajustar a coluna "
        "`Categoria` no Excel (override individual)."
    )


def recategorizar_excel(caminho: Path) -> None:
    """Re-aplica `categorizar()` em todas as linhas da aba `Transações` do
    Excel acumulativo (sem reler PDFs) e reconstrói as abas analíticas.

    Útil quando você editou `categorias.py` ou `categorias_usuario.json`
    e quer propagar as novas regras para o Excel inteiro, sem precisar
    apagar o arquivo e reprocessar todas as faturas. Preserva linhas,
    ordem e cabeçalhos da aba `Transações` — apenas a coluna `Categoria`
    é reescrita.

    ATENÇÃO: edições manuais na coluna `Categoria` do Excel que ainda
    não foram capturadas via `python extrator.py aprender` são
    sobrescritas pelo resultado do dicionário+JSON. Rode `aprender`
    antes se quiser preservá-las.
    """
    if not caminho.exists():
        print(f"Excel não encontrado: {caminho}")
        sys.exit(1)

    df_info, df_transacoes, _ = _carregar_excel_existente(caminho)
    if df_transacoes.empty or "Descrição" not in df_transacoes.columns:
        print(f"Aba 'Transações' vazia ou sem as colunas esperadas em {caminho}.")
        sys.exit(1)

    df_transacoes = df_transacoes.copy()
    categoria_antiga = df_transacoes["Categoria"].fillna("").astype(str)
    descricoes = df_transacoes["Descrição"].fillna("").astype(str)
    categoria_nova = descricoes.map(categorizar)

    mudou = categoria_antiga != categoria_nova
    n_mudancas = int(mudou.sum())
    df_transacoes["Categoria"] = categoria_nova

    salvar_excel_acumulativo(df_info, df_transacoes, caminho)

    print(
        f"Recategorização concluída: {n_mudancas} de {len(df_transacoes)} "
        f"transações tiveram a categoria alterada.\nArquivo: {caminho}"
    )
    if n_mudancas:
        print("\nResumo das mudanças (antiga → nova : qtde):")
        mudancas = pd.DataFrame(
            {"antiga": categoria_antiga[mudou], "nova": categoria_nova[mudou]}
        )
        agg = (
            mudancas.groupby(["antiga", "nova"]).size().reset_index(name="n")
            .sort_values("n", ascending=False)
        )
        for _, row in agg.iterrows():
            print(f"  {row['antiga']:25} → {row['nova']:25} : {int(row['n'])}")

    _imprimir_top_outros_gastos(df_transacoes)
    _imprimir_comparativo_mensal(df_transacoes)


def aprender_do_excel(caminho: Path) -> None:
    """Lê o Excel acumulativo e registra em `categorias_usuario.json` cada
    descrição cuja categoria salva difere do que o dicionário fixo
    devolveria. Útil quando o usuário corrigiu categorias manualmente no
    Excel e quer que essas correções valham para futuras execuções."""
    if not caminho.exists():
        print(f"Excel não encontrado: {caminho}")
        sys.exit(1)

    df = pd.read_excel(caminho, sheet_name="Transações")
    if df.empty or "Descrição" not in df.columns or "Categoria" not in df.columns:
        print(f"Aba 'Transações' vazia ou sem as colunas esperadas em {caminho}.")
        sys.exit(1)

    overrides: dict[str, str] = {}
    for _, linha in df.iterrows():
        desc = str(linha.get("Descrição", "")).strip()
        cat = str(linha.get("Categoria", "")).strip()
        if not desc or not cat:
            continue
        if categorizar_pelo_dicionario(desc) != cat:
            overrides[desc] = cat

    qtd = salvar_categorias_usuario(overrides)
    print(
        f"Aprendizado concluído: {qtd} overrides salvos em "
        f"{CATEGORIAS_USUARIO_ARQUIVO} (lidos de {caminho.name})."
    )
    if qtd == 0:
        print(
            "Nenhuma diferença encontrada entre o dicionário fixo e o "
            "Excel — não há overrides para registrar."
        )


def main() -> None:
    argv = sys.argv[1:]
    if argv and argv[0] in {"aprender", "recategorizar"}:
        comando = argv[0]
        caminho = (
            Path(argv[1]).expanduser().resolve()
            if len(argv) > 1
            else PASTA_SAIDA / ARQUIVO_SAIDA
        )
        if comando == "aprender":
            aprender_do_excel(caminho)
        else:
            recategorizar_excel(caminho)
        return

    alvo = Path(argv[0]).expanduser().resolve() if argv else None
    pdfs = _descobrir_pdfs(alvo)
    if not pdfs:
        sys.exit(1)
    destino = PASTA_SAIDA / ARQUIVO_SAIDA
    PASTA_SAIDA.mkdir(parents=True, exist_ok=True)
    processar(pdfs, destino)
    print("\nConcluído.")


if __name__ == "__main__":
    main()
