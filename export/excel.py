"""Geração da planilha `saida/gastometro.xlsx` a partir do banco.

A lógica de criação das abas, formatação e gráficos vinha de
`extrator.py`. Foi movida pra cá sem alterações funcionais — o
arquivo Excel produzido é idêntico ao formato atual. A diferença é a
fonte: agora os DataFrames vêm do banco (via `db.repository`), não
de uma leitura do próprio Excel anterior.

Função principal: `regenerar_planilha_do_banco(destino)`.

Razão de existir: o Excel deixa de ser fonte de verdade (passa a
ser o banco), mas continua útil como:
  - Backup legível pra leitor humano / planilhas externas.
  - Compatibilidade com o fluxo atual do usuário (abrir o arquivo).
  - Gráficos prontos pro analista que prefere abrir o XLSX.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from db.repository import listar_faturas_df, listar_lancamentos_df
from parsers.base import MES_POR_NUMERO

COLUNAS_INFO = [
    "Arquivo",
    "Banco",
    "Titular",
    "Cartão",
    "Referência",
    "Data de fechamento",
    "Data de vencimento",
    "Valor total (R$)",
    "Qtde. de transações",
]

COLUNAS_TRANSACOES = [
    "Arquivo",
    "Banco",
    "Titular",
    "Cartão",
    "Referência",
    "Data",
    "Descrição",
    "Parcela",
    "Cidade",
    "Valor (R$)",
    "Categoria",
]

MES_POR_NOME = {nome: num for num, nome in MES_POR_NUMERO.items()}

ABAS_COM_FILTRO = {
    "Informações",
    "Transações",
    "Top Comerciantes",
    "Recorrentes",
    "Maiores Gastos",
    "Estornos",
    "Comparativo",
}

ABAS_VALOR_PIVOT = {"Resumo Mensal", "Cartão x Mês", "Cartão x Categoria"}


def _ref_iso_para_nome_br(iso: str | None) -> str:
    """'2026-05' → 'Maio/2026'. Aceita já formatado (passa direto)."""
    if not iso:
        return ""
    if "/" in iso:
        return iso
    if "-" not in iso:
        return iso
    try:
        ano_str, mes_str = iso.split("-", 1)
        ano, mes = int(ano_str), int(mes_str)
    except ValueError:
        return iso
    nome = MES_POR_NUMERO.get(mes, mes_str)
    return f"{nome}/{ano}"


def _data_para_br(valor: object) -> str:
    """`date`/`datetime`/`Timestamp` → `DD/MM/AAAA`. Vazio → ''."""
    if valor is None:
        return ""
    if isinstance(valor, pd.Timestamp):
        if pd.isna(valor):
            return ""
        return valor.strftime("%d/%m/%Y")
    formatador = getattr(valor, "strftime", None)
    if formatador is not None:
        return formatador("%d/%m/%Y")
    return str(valor) if valor else ""


def _banco_de_cartao(cartao: str) -> str:
    """'Banco — Titular' → 'Banco'. Retrocompatibilidade."""
    if not cartao:
        return ""
    return cartao.split(" — ", 1)[0].strip()


def _identificador_cartao(banco: str, titular: str) -> str:
    """Espelha `extrator._identificador_cartao` (fonte única no repository)."""
    banco = (banco or "").strip()
    titular = (titular or "").strip()
    if not banco and not titular:
        return ""
    if not titular:
        return banco
    if not banco:
        return titular
    return f"{banco} — {titular}"


def _df_faturas_para_info(df_fat: pd.DataFrame) -> pd.DataFrame:
    """Traduz schema do banco → schema da aba `Informações` (CLI atual)."""
    if df_fat.empty:
        return pd.DataFrame(columns=COLUNAS_INFO)
    registros: list[dict[str, object]] = []
    for _, linha in df_fat.iterrows():
        cartao = str(linha.get("conta", "") or "")
        registros.append(
            {
                "Arquivo": linha.get("arquivo", ""),
                "Banco": _banco_de_cartao(cartao),
                "Titular": linha.get("pessoa", "") or "",
                "Cartão": cartao,
                "Referência": _ref_iso_para_nome_br(linha.get("referencia_mes")),
                "Data de fechamento": _data_para_br(linha.get("fechamento")),
                "Data de vencimento": _data_para_br(linha.get("vencimento")),
                "Valor total (R$)": float(linha.get("valor_total_declarado") or 0.0),
                "Qtde. de transações": int(linha.get("qtde_transacoes") or 0),
            }
        )
    return pd.DataFrame(registros, columns=COLUNAS_INFO)


def _df_lancamentos_para_transacoes(df_lanc: pd.DataFrame) -> pd.DataFrame:
    """Traduz schema do banco → schema da aba `Transações` (CLI atual)."""
    if df_lanc.empty:
        return pd.DataFrame(columns=COLUNAS_TRANSACOES)
    registros: list[dict[str, object]] = []
    for _, linha in df_lanc.iterrows():
        cartao = str(linha.get("conta", "") or "")
        registros.append(
            {
                "Arquivo": linha.get("arquivo", "") or "",
                "Banco": _banco_de_cartao(cartao),
                "Titular": linha.get("pessoa", "") or "",
                "Cartão": cartao,
                "Referência": _ref_iso_para_nome_br(linha.get("referencia_mes")),
                "Data": _data_para_br(linha.get("data")),
                "Descrição": linha.get("descricao", "") or "",
                "Parcela": linha.get("parcela", "") or "",
                "Cidade": linha.get("cidade", "") or "",
                "Valor (R$)": float(linha.get("valor", 0.0) or 0.0),
                "Categoria": linha.get("categoria", "") or "",
            }
        )
    return pd.DataFrame(registros, columns=COLUNAS_TRANSACOES)


def _ordenar_transacoes(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    df = df.copy()
    df["_data_ord"] = pd.to_datetime(df["Data"], format="%d/%m/%Y", errors="coerce")
    df = df.sort_values(["_data_ord", "Arquivo"], kind="stable").reset_index(drop=True)
    return df.drop(columns="_data_ord")


def _chave_ordenacao_referencia(referencia: str) -> tuple[int, int]:
    """Converte 'Maio/2026' em (2026, 5) para ordenar cronologicamente."""
    try:
        nome_mes, ano = referencia.split("/")
        return (int(ano), MES_POR_NOME.get(nome_mes, 0))
    except (ValueError, AttributeError):
        return (0, 0)


def _construir_resumo_geral(df_transacoes: pd.DataFrame) -> pd.DataFrame:
    if df_transacoes.empty:
        return pd.DataFrame(columns=["Categoria", "Valor (R$)"])
    resumo = (
        df_transacoes.groupby("Categoria", as_index=False)["Valor (R$)"]
        .sum()
        .sort_values("Valor (R$)", ascending=False)
        .reset_index(drop=True)
    )
    total = pd.DataFrame(
        [{"Categoria": "TOTAL GERAL", "Valor (R$)": resumo["Valor (R$)"].sum()}]
    )
    return pd.concat([resumo, total], ignore_index=True)


def _construir_resumo_mensal(df_transacoes: pd.DataFrame) -> pd.DataFrame:
    if df_transacoes.empty or "Referência" not in df_transacoes.columns:
        return pd.DataFrame()
    pivot = pd.pivot_table(
        df_transacoes,
        index="Referência",
        columns="Categoria",
        values="Valor (R$)",
        aggfunc="sum",
        fill_value=0.0,
    )
    referencias_ordenadas = sorted(pivot.index, key=_chave_ordenacao_referencia)
    pivot = pivot.loc[referencias_ordenadas]
    pivot["Total"] = pivot.sum(axis=1)

    totais = pivot["Total"].tolist()
    variacoes: list[float | None] = []
    for i, atual in enumerate(totais):
        if i == 0:
            variacoes.append(None)
            continue
        anterior = totais[i - 1]
        if anterior == 0:
            variacoes.append(None)
        else:
            variacoes.append((atual - anterior) / anterior * 100.0)
    pivot["Variação %"] = variacoes

    soma_linhas = pivot.drop(columns=["Variação %"]).sum(axis=0)
    pivot.loc["TOTAL"] = soma_linhas
    pivot.at["TOTAL", "Variação %"] = None
    return pivot.reset_index()


def _construir_cartao_x_mes(df_transacoes: pd.DataFrame) -> pd.DataFrame:
    if df_transacoes.empty:
        return pd.DataFrame()
    if "Referência" not in df_transacoes.columns or "Cartão" not in df_transacoes.columns:
        return pd.DataFrame()
    pivot = pd.pivot_table(
        df_transacoes,
        index="Referência",
        columns="Cartão",
        values="Valor (R$)",
        aggfunc="sum",
        fill_value=0.0,
    )
    referencias_ordenadas = sorted(pivot.index, key=_chave_ordenacao_referencia)
    pivot = pivot.loc[referencias_ordenadas]
    pivot["Total"] = pivot.sum(axis=1)
    pivot.loc["TOTAL"] = pivot.sum(axis=0)
    return pivot.reset_index()


def _construir_cartao_x_categoria(df_transacoes: pd.DataFrame) -> pd.DataFrame:
    if df_transacoes.empty:
        return pd.DataFrame()
    if "Categoria" not in df_transacoes.columns or "Cartão" not in df_transacoes.columns:
        return pd.DataFrame()
    pivot = pd.pivot_table(
        df_transacoes,
        index="Categoria",
        columns="Cartão",
        values="Valor (R$)",
        aggfunc="sum",
        fill_value=0.0,
    )
    pivot["Total"] = pivot.sum(axis=1)
    pivot = pivot.sort_values("Total", ascending=False)
    pivot.loc["TOTAL"] = pivot.sum(axis=0)
    return pivot.reset_index()


def _construir_resumo_por_cartao(
    df_info: pd.DataFrame, df_transacoes: pd.DataFrame
) -> pd.DataFrame:
    if df_transacoes.empty or "Cartão" not in df_transacoes.columns:
        return pd.DataFrame()

    df_t = df_transacoes.copy()
    linhas: list[dict[str, object]] = []
    for cartao, grupo in df_t.groupby("Cartão", sort=False):
        if not cartao:
            continue
        banco = str(grupo["Banco"].iloc[0]) if "Banco" in grupo.columns else ""
        titular = str(grupo["Titular"].iloc[0]) if "Titular" in grupo.columns else ""
        refs = sorted(set(grupo["Referência"]), key=_chave_ordenacao_referencia)
        total = float(grupo["Valor (R$)"].sum())
        qtd_tx = len(grupo)
        qtd_faturas = grupo["Arquivo"].nunique() if "Arquivo" in grupo.columns else len(refs)
        media_fatura = total / qtd_faturas if qtd_faturas else 0.0
        ticket = total / qtd_tx if qtd_tx else 0.0
        linhas.append(
            {
                "Cartão": cartao,
                "Banco": banco,
                "Titular": titular,
                "Qtde. Faturas": qtd_faturas,
                "Qtde. Transações": qtd_tx,
                "Primeira Referência": refs[0] if refs else "",
                "Última Referência": refs[-1] if refs else "",
                "Total Gasto (R$)": total,
                "Média por Fatura (R$)": media_fatura,
                "Ticket Médio (R$)": ticket,
            }
        )

    if not linhas:
        return pd.DataFrame()

    df_resumo = pd.DataFrame(linhas).sort_values(
        "Total Gasto (R$)", ascending=False
    )

    total_geral = {
        "Cartão": "TOTAL",
        "Banco": "",
        "Titular": "",
        "Qtde. Faturas": int(df_resumo["Qtde. Faturas"].sum()),
        "Qtde. Transações": int(df_resumo["Qtde. Transações"].sum()),
        "Primeira Referência": "",
        "Última Referência": "",
        "Total Gasto (R$)": float(df_resumo["Total Gasto (R$)"].sum()),
        "Média por Fatura (R$)": (
            float(df_resumo["Total Gasto (R$)"].sum())
            / int(df_resumo["Qtde. Faturas"].sum())
            if df_resumo["Qtde. Faturas"].sum()
            else 0.0
        ),
        "Ticket Médio (R$)": (
            float(df_resumo["Total Gasto (R$)"].sum())
            / int(df_resumo["Qtde. Transações"].sum())
            if df_resumo["Qtde. Transações"].sum()
            else 0.0
        ),
    }
    return pd.concat(
        [df_resumo, pd.DataFrame([total_geral])], ignore_index=True
    )


def _construir_estornos(df_transacoes: pd.DataFrame) -> pd.DataFrame:
    if df_transacoes.empty or "Valor (R$)" not in df_transacoes.columns:
        return pd.DataFrame()

    estornos = df_transacoes[df_transacoes["Valor (R$)"] < 0].copy()
    if estornos.empty:
        return pd.DataFrame()

    colunas = [
        "Data",
        "Referência",
        "Descrição",
        "Categoria",
        "Cartão",
        "Cidade",
        "Valor (R$)",
        "Arquivo",
    ]
    colunas = [c for c in colunas if c in estornos.columns]
    estornos["_abs"] = estornos["Valor (R$)"].abs()
    estornos["_ref_ord"] = estornos["Referência"].apply(_chave_ordenacao_referencia)
    estornos = estornos.sort_values(
        ["_ref_ord", "_abs"], ascending=[False, False]
    )
    return estornos[colunas].reset_index(drop=True)


def _construir_comparativo(df_transacoes: pd.DataFrame) -> pd.DataFrame:
    if df_transacoes.empty or "Referência" not in df_transacoes.columns:
        return pd.DataFrame()
    if "Categoria" not in df_transacoes.columns:
        return pd.DataFrame()

    refs = sorted(
        {str(r) for r in df_transacoes["Referência"] if str(r)},
        key=_chave_ordenacao_referencia,
    )
    if len(refs) < 2:
        return pd.DataFrame()
    ultimo, anterior = refs[-1], refs[-2]

    def soma_por_cat(ref: str) -> dict[str, float]:
        sub = df_transacoes[df_transacoes["Referência"] == ref]
        return {
            str(k): float(v)
            for k, v in sub.groupby("Categoria")["Valor (R$)"].sum().items()
        }

    cats_ult = soma_por_cat(ultimo)
    cats_ant = soma_por_cat(anterior)
    todas = sorted(set(cats_ult) | set(cats_ant))

    linhas: list[dict[str, object]] = []
    for cat in todas:
        atual = cats_ult.get(cat, 0.0)
        ant = cats_ant.get(cat, 0.0)
        diff = atual - ant
        var = (diff / ant * 100.0) if ant else None
        linhas.append(
            {
                "Categoria": cat,
                f"{anterior} (R$)": ant,
                f"{ultimo} (R$)": atual,
                "Δ Absoluto (R$)": diff,
                "Δ %": var,
            }
        )

    if not linhas:
        return pd.DataFrame()

    df = pd.DataFrame(linhas)
    df["_abs"] = df["Δ Absoluto (R$)"].abs()
    df = df.sort_values("_abs", ascending=False).drop(columns="_abs")

    total_ant = float(sum(cats_ant.values()))
    total_ult = float(sum(cats_ult.values()))
    total_diff = total_ult - total_ant
    total_var = (total_diff / total_ant * 100.0) if total_ant else None
    df = pd.concat(
        [
            df,
            pd.DataFrame(
                [
                    {
                        "Categoria": "TOTAL",
                        f"{anterior} (R$)": total_ant,
                        f"{ultimo} (R$)": total_ult,
                        "Δ Absoluto (R$)": total_diff,
                        "Δ %": total_var,
                    }
                ]
            ),
        ],
        ignore_index=True,
    )
    return df


def _construir_top_transacoes(
    df_transacoes: pd.DataFrame, top_n: int = 20
) -> pd.DataFrame:
    if df_transacoes.empty:
        return pd.DataFrame()

    gastos = df_transacoes[df_transacoes["Valor (R$)"] > 0].copy()
    if gastos.empty:
        return pd.DataFrame()

    colunas = [
        "Data",
        "Referência",
        "Descrição",
        "Categoria",
        "Cartão",
        "Parcela",
        "Cidade",
        "Valor (R$)",
    ]
    colunas = [c for c in colunas if c in gastos.columns]
    return (
        gastos.sort_values("Valor (R$)", ascending=False)
        .head(top_n)[colunas]
        .reset_index(drop=True)
    )


def _construir_top_comerciantes(
    df_transacoes: pd.DataFrame, top_n: int = 30
) -> pd.DataFrame:
    if df_transacoes.empty or "Descrição" not in df_transacoes.columns:
        return pd.DataFrame()

    gastos = df_transacoes[df_transacoes["Valor (R$)"] > 0].copy()
    if gastos.empty:
        return pd.DataFrame()

    agg = (
        gastos.groupby("Descrição")
        .agg(
            **{
                "Qtde.": ("Valor (R$)", "count"),
                "Total (R$)": ("Valor (R$)", "sum"),
                "Ticket Médio (R$)": ("Valor (R$)", "mean"),
                "Categoria": ("Categoria", lambda s: s.mode().iat[0] if not s.mode().empty else ""),
                "Cartão(ões)": ("Cartão", lambda s: ", ".join(sorted({str(v) for v in s if str(v).strip()}))),
            }
        )
        .reset_index()
        .sort_values("Total (R$)", ascending=False)
        .head(top_n)
        .reset_index(drop=True)
    )
    return agg[
        [
            "Descrição",
            "Categoria",
            "Cartão(ões)",
            "Qtde.",
            "Total (R$)",
            "Ticket Médio (R$)",
        ]
    ]


def _construir_recorrentes(
    df_transacoes: pd.DataFrame, meses_min: int = 3
) -> pd.DataFrame:
    if df_transacoes.empty or "Descrição" not in df_transacoes.columns:
        return pd.DataFrame()
    if "Referência" not in df_transacoes.columns:
        return pd.DataFrame()

    gastos = df_transacoes[df_transacoes["Valor (R$)"] > 0].copy()
    if gastos.empty:
        return pd.DataFrame()

    agg = (
        gastos.groupby("Descrição")
        .agg(
            **{
                "Meses": ("Referência", "nunique"),
                "Qtde. Transações": ("Valor (R$)", "count"),
                "Total (R$)": ("Valor (R$)", "sum"),
                "Média Mensal (R$)": ("Valor (R$)", "sum"),
                "Categoria": ("Categoria", lambda s: s.mode().iat[0] if not s.mode().empty else ""),
                "Cartão(ões)": ("Cartão", lambda s: ", ".join(sorted({str(v) for v in s if str(v).strip()}))),
            }
        )
        .reset_index()
    )
    agg = agg[agg["Meses"] >= meses_min].copy()
    if agg.empty:
        return pd.DataFrame()
    agg["Média Mensal (R$)"] = agg["Total (R$)"] / agg["Meses"]
    agg = agg.sort_values(
        ["Total (R$)", "Meses"], ascending=[False, False]
    ).reset_index(drop=True)
    return agg[
        [
            "Descrição",
            "Categoria",
            "Cartão(ões)",
            "Meses",
            "Qtde. Transações",
            "Total (R$)",
            "Média Mensal (R$)",
        ]
    ]


def _formatar_planilha(writer: object, nome_aba: str, df: pd.DataFrame) -> None:
    aba = writer.sheets[nome_aba]  # type: ignore[attr-defined]

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx in range(1, len(df.columns) + 1):
        cell = aba.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    colunas_percent_idx = [
        df.columns.get_loc(c) + 1 for c in df.columns if "%" in str(c)
    ]
    if nome_aba in ABAS_VALOR_PIVOT:
        colunas_valor_idx = [
            i for i in range(2, len(df.columns) + 1) if i not in colunas_percent_idx
        ]
    else:
        colunas_valor_idx = [
            df.columns.get_loc(c) + 1
            for c in df.columns
            if (
                "Valor" in c
                or "Total" in c
                or "Médio" in c
                or "Média" in c
                or "R$" in c
                or c == "Ticket"
            )
            and "%" not in str(c)
        ]
    for col_idx in colunas_valor_idx:
        for row in range(2, len(df) + 2):
            celula = aba.cell(row=row, column=col_idx)
            if isinstance(celula.value, (int, float)):
                celula.number_format = "R$ #,##0.00"
    for col_idx in colunas_percent_idx:
        for row in range(2, len(df) + 2):
            celula = aba.cell(row=row, column=col_idx)
            if isinstance(celula.value, (int, float)):
                celula.number_format = '+0.0"%";-0.0"%";0.0"%"'

    for col_idx, coluna in enumerate(df.columns, start=1):
        valores = [str(v) for v in df[coluna].astype(str).tolist()] + [str(coluna)]
        largura = min(max(len(v) for v in valores) + 2, 60)
        aba.column_dimensions[get_column_letter(col_idx)].width = largura

    if nome_aba in ABAS_COM_FILTRO and len(df) > 0 and len(df.columns) > 0:
        ultima_col = get_column_letter(len(df.columns))
        aba.auto_filter.ref = f"A1:{ultima_col}{len(df) + 1}"

    aba.freeze_panes = "A2"


def _adicionar_grafico_categorias(writer: object, df_resumo: pd.DataFrame) -> None:
    if df_resumo.empty or len(df_resumo) < 3:
        return
    aba = writer.sheets.get("Resumo por Categoria")  # type: ignore[attr-defined]
    if aba is None:
        return
    ultima_linha_dados = len(df_resumo)
    chart = PieChart()
    chart.title = "Distribuição por Categoria"
    chart.height = 12
    chart.width = 18
    rotulos = Reference(aba, min_col=1, min_row=2, max_row=ultima_linha_dados)
    dados = Reference(
        aba, min_col=2, min_row=1, max_row=ultima_linha_dados
    )
    chart.add_data(dados, titles_from_data=True)
    chart.set_categories(rotulos)
    aba.add_chart(chart, "D2")


def _adicionar_grafico_tendencia_cartoes(
    writer: object, df_cartao_mes: pd.DataFrame
) -> None:
    if df_cartao_mes.empty or "Total" not in df_cartao_mes.columns:
        return
    aba = writer.sheets.get("Cartão x Mês")  # type: ignore[attr-defined]
    if aba is None:
        return
    ultima_linha_dados = len(df_cartao_mes)
    n_meses = ultima_linha_dados - 1
    if n_meses < 2:
        return
    col_total = df_cartao_mes.columns.get_loc("Total") + 1
    if col_total < 3:
        return

    chart = LineChart()
    chart.title = "Tendência por Cartão"
    chart.y_axis.title = "R$"
    chart.x_axis.title = "Mês"
    chart.height = 12
    chart.width = 24
    dados = Reference(
        aba,
        min_col=2,
        max_col=col_total - 1,
        min_row=1,
        max_row=1 + n_meses,
    )
    categorias = Reference(aba, min_col=1, min_row=2, max_row=1 + n_meses)
    chart.add_data(dados, titles_from_data=True)
    chart.set_categories(categorias)
    aba.add_chart(chart, f"A{ultima_linha_dados + 3}")


def _adicionar_grafico_mensal(writer: object, df_mensal: pd.DataFrame) -> None:
    if df_mensal.empty or "Total" not in df_mensal.columns:
        return
    aba = writer.sheets.get("Resumo Mensal")  # type: ignore[attr-defined]
    if aba is None:
        return
    ultima_linha_dados = len(df_mensal)
    n_meses = ultima_linha_dados - 1
    if n_meses < 2:
        return
    col_total = df_mensal.columns.get_loc("Total") + 1

    chart = BarChart()
    chart.type = "col"
    chart.title = "Total Mensal"
    chart.y_axis.title = "R$"
    chart.x_axis.title = "Mês"
    chart.height = 10
    chart.width = 22
    dados = Reference(aba, min_col=col_total, min_row=1, max_row=1 + n_meses)
    categorias = Reference(aba, min_col=1, min_row=2, max_row=1 + n_meses)
    chart.add_data(dados, titles_from_data=True)
    chart.set_categories(categorias)
    aba.add_chart(chart, f"A{ultima_linha_dados + 3}")


def salvar_planilha(
    df_info: pd.DataFrame, df_transacoes: pd.DataFrame, destino: Path
) -> None:
    """Escreve o XLSX completo (todas as abas e gráficos) em `destino`.

    Function pública pra quem já tem os DataFrames no schema do Excel
    em mãos (legado). Pra alimentar a partir do banco, use
    `regenerar_planilha_do_banco`.
    """
    df_transacoes = _ordenar_transacoes(df_transacoes)
    df_info = df_info.reindex(columns=COLUNAS_INFO)

    df_resumo = _construir_resumo_geral(df_transacoes)
    df_mensal = _construir_resumo_mensal(df_transacoes)
    df_comparativo = _construir_comparativo(df_transacoes)
    df_cartao_mes = _construir_cartao_x_mes(df_transacoes)
    df_cartao_cat = _construir_cartao_x_categoria(df_transacoes)
    df_resumo_cartao = _construir_resumo_por_cartao(df_info, df_transacoes)
    df_maiores = _construir_top_transacoes(df_transacoes)
    df_estornos = _construir_estornos(df_transacoes)
    df_top = _construir_top_comerciantes(df_transacoes)
    df_recorrentes = _construir_recorrentes(df_transacoes)

    destino.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(destino, engine="openpyxl") as writer:
        df_info.to_excel(writer, sheet_name="Informações", index=False)
        df_transacoes.to_excel(writer, sheet_name="Transações", index=False)
        df_resumo.to_excel(writer, sheet_name="Resumo por Categoria", index=False)
        if not df_mensal.empty:
            df_mensal.to_excel(writer, sheet_name="Resumo Mensal", index=False)
        if not df_comparativo.empty:
            df_comparativo.to_excel(writer, sheet_name="Comparativo", index=False)
        if not df_resumo_cartao.empty:
            df_resumo_cartao.to_excel(
                writer, sheet_name="Resumo por Cartão", index=False
            )
        if not df_cartao_mes.empty:
            df_cartao_mes.to_excel(writer, sheet_name="Cartão x Mês", index=False)
        if not df_cartao_cat.empty:
            df_cartao_cat.to_excel(
                writer, sheet_name="Cartão x Categoria", index=False
            )
        if not df_maiores.empty:
            df_maiores.to_excel(writer, sheet_name="Maiores Gastos", index=False)
        if not df_estornos.empty:
            df_estornos.to_excel(writer, sheet_name="Estornos", index=False)
        if not df_top.empty:
            df_top.to_excel(writer, sheet_name="Top Comerciantes", index=False)
        if not df_recorrentes.empty:
            df_recorrentes.to_excel(writer, sheet_name="Recorrentes", index=False)

        _formatar_planilha(writer, "Informações", df_info)
        _formatar_planilha(writer, "Transações", df_transacoes)
        _formatar_planilha(writer, "Resumo por Categoria", df_resumo)
        if not df_mensal.empty:
            _formatar_planilha(writer, "Resumo Mensal", df_mensal)
        if not df_comparativo.empty:
            _formatar_planilha(writer, "Comparativo", df_comparativo)
        if not df_resumo_cartao.empty:
            _formatar_planilha(writer, "Resumo por Cartão", df_resumo_cartao)
        if not df_cartao_mes.empty:
            _formatar_planilha(writer, "Cartão x Mês", df_cartao_mes)
        if not df_cartao_cat.empty:
            _formatar_planilha(writer, "Cartão x Categoria", df_cartao_cat)
        if not df_maiores.empty:
            _formatar_planilha(writer, "Maiores Gastos", df_maiores)
        if not df_estornos.empty:
            _formatar_planilha(writer, "Estornos", df_estornos)
        if not df_top.empty:
            _formatar_planilha(writer, "Top Comerciantes", df_top)
        if not df_recorrentes.empty:
            _formatar_planilha(writer, "Recorrentes", df_recorrentes)

        _adicionar_grafico_categorias(writer, df_resumo)
        if not df_mensal.empty:
            _adicionar_grafico_mensal(writer, df_mensal)
        if not df_cartao_mes.empty:
            _adicionar_grafico_tendencia_cartoes(writer, df_cartao_mes)


def regenerar_planilha_do_banco(destino: Path) -> tuple[int, int]:
    """Lê banco, traduz schema, escreve `destino`. Devolve (faturas, lancamentos).

    Pra Fase 1+: o Excel é um espelho regenerável a qualquer momento.
    Edições devem acontecer no banco (Streamlit a partir da Fase 2);
    o XLSX é só visualização/backup legível.
    """
    df_fat = listar_faturas_df()
    df_lanc = listar_lancamentos_df(sinal_estorno_negativo=True)

    df_info = _df_faturas_para_info(df_fat)
    df_transacoes = _df_lancamentos_para_transacoes(df_lanc)

    salvar_planilha(df_info, df_transacoes, destino)
    return len(df_info), len(df_transacoes)
